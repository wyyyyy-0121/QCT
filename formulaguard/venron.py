"""Version-order and formula-transition primitives for the frozen VEnron V0 gate."""

from __future__ import annotations

import hashlib
import re
from collections import Counter
from pathlib import Path, PurePosixPath
from typing import Iterable, Mapping, Sequence

import openpyxl


ORDER_MEMBER = "VEnron1.0/Version/FileOrder.xls"
GROUP_PATTERN = re.compile(r"^(\d+)_(\d+)_(.+)$")
MD5_TOKEN_PATTERN = re.compile(r"^[0-9a-f]{29,32}$")


def _integer(value: object) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def normalize_order_path(value: object, group_name: str) -> str:
    if not isinstance(value, str) or not value:
        raise ValueError("VEnron order path is not a string")
    candidate = value.replace("\\", "/")
    if not candidate.startswith("../"):
        raise ValueError(f"VEnron order path is not parent-relative: {value!r}")
    parts = candidate[3:].split("/")
    if len(parts) != 2 or any(part in {"", ".", ".."} for part in parts):
        raise ValueError(f"VEnron order path has an unsafe shape: {value!r}")
    if parts[0] != group_name or not parts[1].lower().endswith(".xls"):
        raise ValueError(f"VEnron order path/group mismatch: {value!r}")
    relative = PurePosixPath("VEnron1.0", *parts)
    return relative.as_posix()


def parse_order_workbook(
    path: Path,
    member_paths: Iterable[str],
    *,
    expected_groups: int = 360,
    expected_workbooks: int = 7_294,
) -> list[dict[str, object]]:
    expected_members = {
        value for value in member_paths if value != ORDER_MEMBER and value.lower().endswith(".xls")
    }
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    records: list[dict[str, object]] = []
    seen_paths: set[str] = set()
    seen_group_ids: set[int] = set()
    try:
        for sheet in workbook.worksheets:
            match = GROUP_PATTERN.fullmatch(sheet.title)
            if not match:
                if sheet.title != "Group List":
                    raise ValueError(f"unexpected VEnron order sheet: {sheet.title!r}")
                continue
            group_id = int(match.group(1))
            declared_count = int(match.group(2))
            seen_group_ids.add(group_id)
            group_rows: list[dict[str, object]] = []
            # Only A, E, and F are retained. Interleaved email rows have no file path in F.
            for row in sheet.iter_rows(min_col=1, max_col=6, values_only=True):
                order = _integer(row[0])
                md5_token = str(row[4]).lower() if row[4] is not None else ""
                file_path = row[5]
                if order is None or not isinstance(file_path, str) or not file_path.lower().endswith(".xls"):
                    continue
                if not MD5_TOKEN_PATTERN.fullmatch(md5_token):
                    raise ValueError(f"invalid VEnron source MD5 in group {group_id}")
                md5 = md5_token.zfill(32)
                relative = normalize_order_path(file_path, sheet.title)
                if relative not in expected_members or relative in seen_paths:
                    raise ValueError(f"VEnron order path is absent or duplicated: {relative!r}")
                if not PurePosixPath(relative).name.lower().endswith(f"_{md5_token}.xls"):
                    raise ValueError(f"VEnron order path/MD5 mismatch: {relative!r}")
                seen_paths.add(relative)
                group_rows.append({
                    "group_id": group_id,
                    "group_name": sheet.title,
                    "version_order": order,
                    "source_md5": md5,
                    "source_md5_token": md5_token,
                    "source_relative_path": relative,
                })
            group_rows.sort(key=lambda row: int(row["version_order"]))
            if len(group_rows) != declared_count or [
                int(row["version_order"]) for row in group_rows
            ] != list(range(1, declared_count + 1)):
                raise ValueError(f"VEnron version order is incomplete for group {group_id}")
            records.extend(group_rows)
    finally:
        workbook.close()

    if seen_group_ids != set(range(1, expected_groups + 1)):
        raise ValueError("VEnron order workbook group IDs are incomplete")
    if len(records) != expected_workbooks or seen_paths != expected_members:
        raise ValueError("VEnron order workbook/member accounting is incomplete")
    return records


def inspect_formula_workbook(path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    formulas: list[dict[str, str]] = []
    sheet_titles: list[str] = []
    try:
        for sheet in workbook.worksheets:
            sheet_titles.append(sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        formulas.append({
                            "sheet": sheet.title,
                            "address": cell.coordinate,
                            "formula": str(cell.value).strip(),
                        })
    finally:
        workbook.close()
    formulas.sort(key=lambda row: (row["sheet"], row["address"]))
    if len({(row["sheet"], row["address"]) for row in formulas}) != len(formulas):
        raise ValueError("duplicate formula key in converted VEnron workbook")
    return {
        "sheet_count": len(sheet_titles),
        "sheet_titles": sheet_titles,
        "formula_count": len(formulas),
        "formulas": formulas,
    }


def _formula_map(profile: Mapping[str, object]) -> dict[tuple[str, str], str]:
    rows = profile.get("formulas")
    if not isinstance(rows, list):
        raise ValueError("VEnron profile has no formula list")
    result: dict[tuple[str, str], str] = {}
    for row in rows:
        if not isinstance(row, Mapping):
            raise ValueError("VEnron formula profile row is malformed")
        key = (str(row.get("sheet", "")), str(row.get("address", "")))
        formula = str(row.get("formula", "")).strip()
        if not all(key) or not formula or key in result:
            raise ValueError("VEnron formula profile identity is invalid")
        result[key] = formula
    return result


def compare_formula_profiles(
    previous: Mapping[str, object],
    current: Mapping[str, object],
) -> dict[str, object]:
    left = _formula_map(previous)
    right = _formula_map(current)
    shared = set(left) & set(right)
    changed = {key for key in shared if left[key] != right[key]}
    added = set(right) - set(left)
    removed = set(left) - set(right)
    unchanged = shared - changed
    direct_fraction = len(changed) / len(shared) if shared else 0.0
    bulk_direct = len(changed) >= 20 and direct_fraction >= 0.50
    bulk_add_remove = len(added) + len(removed) >= 20
    address_only_move = bool(added and removed) and not changed and (
        Counter(right[key] for key in added) == Counter(left[key] for key in removed)
    )
    return {
        "previous_formula_count": len(left),
        "current_formula_count": len(right),
        "shared_formula_keys": len(shared),
        "unchanged_formula_keys": len(unchanged),
        "direct_formula_text_changes": len(changed),
        "formula_additions": len(added),
        "formula_removals": len(removed),
        "direct_change_fraction_of_shared": round(direct_fraction, 12),
        "has_direct_formula_text_change": bool(changed),
        "single_direct": len(changed) == 1,
        "multi_direct": len(changed) >= 2,
        "bulk_direct_rewrite": bulk_direct,
        "bulk_add_remove": bulk_add_remove,
        "address_only_formula_move": address_only_move,
        "no_formula_text_change": not changed and not added and not removed,
        "nonbulk_multi_direct": len(changed) >= 2 and not bulk_direct and not bulk_add_remove,
    }


def stable_record_id(*parts: object) -> str:
    return hashlib.sha256("\0".join(str(part) for part in parts).encode("utf-8")).hexdigest()
