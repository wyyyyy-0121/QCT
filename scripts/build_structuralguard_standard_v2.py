"""Build the corrected, same-source StructuralGuard development benchmark v2."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import zipfile
from copy import deepcopy
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

PROTOCOL = "structuralguard_standard_benchmark_v2"
FIXED_TIME = datetime(2000, 1, 1, tzinfo=UTC)
BLOCK_SPECS = {
    "retail_sales": ("G37", range(28, 48)),
    "inventory": ("F40", range(31, 51)),
    "project_budget": ("F34", range(25, 45)),
    "student_grades": ("G25", range(16, 36)),
    "saas_revenue": ("F45", range(36, 56)),
}


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def normalized_formula(value: str) -> str:
    return "".join(value.split()).upper()


def translated(formula: str, source_cell: str, target_cell: str) -> str:
    return Translator(formula, origin=source_cell).translate_formula(target_cell)


def canonical_core_xml(value: bytes) -> bytes:
    namespaces = {
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
        "dc": "http://purl.org/dc/elements/1.1/",
        "dcterms": "http://purl.org/dc/terms/",
        "xsi": "http://www.w3.org/2001/XMLSchema-instance",
    }
    for prefix, uri in namespaces.items():
        ElementTree.register_namespace(prefix, uri)
    root = ElementTree.fromstring(value)
    modified = root.find("{http://purl.org/dc/terms/}modified")
    if modified is not None:
        modified.text = "2000-01-01T00:00:00Z"
    return ElementTree.tostring(root, encoding="utf-8")


def save_case(wb, path: Path) -> str:
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.properties.lastPrinted = None
    if getattr(wb, "calculation", None) is not None:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(buffer, "r") as source, zipfile.ZipFile(
        path,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            content = source.read(name)
            if name == "docProps/core.xml":
                content = canonical_core_xml(content)
            target.writestr(info, content)
    return sha256_file(path)


def formula_count(wb) -> int:
    return sum(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
    )


def copy_case(v1: Path, output: Path, source_case: dict, case_id: str) -> dict:
    source = v1 / source_case["workbook"]
    target = output / "public" / "workbooks" / f"{case_id}.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, target)
    case = deepcopy(source_case)
    case["case_id"] = case_id
    case["workbook"] = f"public/workbooks/{case_id}.xlsx"
    case["sha256"] = sha256_file(target)
    case["derived_from_case"] = source_case["case_id"]
    return case


def corrected_block_case(
    v1: Path,
    output: Path,
    clean_case: dict,
    singleton_case: dict,
    case_id: str,
) -> dict:
    scenario = str(clean_case["scenario"])
    source_cell, rows = BLOCK_SPECS[scenario]
    source_item = next(
        item for item in singleton_case["errors"] if str(item["cell"]) == source_cell
    )
    sheet = str(source_item["sheet"])
    column_match = re.match(r"[A-Z]+", source_cell)
    if column_match is None:
        raise ValueError(f"invalid source cell: {source_cell}")
    column = column_match.group()
    wb = load_workbook(v1 / clean_case["workbook"], data_only=False)
    errors = []
    for row in rows:
        target_cell = f"{column}{row}"
        expected = translated(str(source_item["expected_formula"]), source_cell, target_cell)
        clean_formula = wb[sheet][target_cell].value
        if not isinstance(clean_formula, str) or normalized_formula(clean_formula) != normalized_formula(expected):
            raise ValueError(
                f"translated repair does not match clean formula: {scenario} {sheet}!{target_cell}"
            )
        injected = translated(str(source_item["injected_formula"]), source_cell, target_cell)
        if normalized_formula(injected) == normalized_formula(expected):
            raise ValueError(f"mutation is formula-equivalent: {scenario} {sheet}!{target_cell}")
        wb[sheet][target_cell].value = injected
        errors.append({
            "sheet": sheet,
            "cell": target_cell,
            "category": str(source_item["category"]),
            "injected_formula": injected,
            "expected_formula": expected,
            "source_template_cell": source_cell,
            "mutation_condition": "coherent_block",
        })
    target = output / "public" / "workbooks" / f"{case_id}.xlsx"
    digest = save_case(wb, target)
    return {
        "case_id": case_id,
        "condition": "coherent_block",
        "scenario": scenario,
        "source_file": clean_case["source_file"],
        "workbook": f"public/workbooks/{case_id}.xlsx",
        "sha256": digest,
        "formula_cells": formula_count(wb),
        "errors": errors,
        "source_sha256": clean_case["source_sha256"],
        "source_seed": clean_case["source_seed"],
        "derived_from_case": clean_case["case_id"],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("v1_dataset", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()
    v1 = args.v1_dataset.resolve()
    output = args.output_dir.resolve()
    labels = json.loads((v1 / "labels" / "answer_keys.json").read_text(encoding="utf-8"))
    if labels.get("protocol") != "structuralguard_standard_benchmark_v1":
        raise SystemExit("v2 must be derived from StructuralGuard standard benchmark v1")
    for name in ("public", "labels", "metadata.json", "SHA256SUMS.txt"):
        target = output / name
        if target.is_dir():
            shutil.rmtree(target)
        elif target.exists():
            target.unlink()
    (output / "public" / "workbooks").mkdir(parents=True, exist_ok=True)
    (output / "labels").mkdir(parents=True, exist_ok=True)
    canonical_readme = (
        Path(__file__).resolve().parents[1]
        / "data"
        / "structuralguard_standard_v2"
        / "README.md"
    )
    if not (output / "README.md").exists() and canonical_readme.exists():
        shutil.copyfile(canonical_readme, output / "README.md")

    by_scenario_condition = {
        (str(case["scenario"]), str(case["condition"])): case
        for case in labels["cases"]
    }
    scenario_order = [
        str(case["scenario"]) for case in labels["cases"] if case["condition"] == "clean"
    ]
    cases = []
    for index, scenario in enumerate(scenario_order, 1):
        for condition in ("singleton", "clean", "coherent_block", "systematic_column"):
            case_id = f"sgv2_{condition}_{index:02d}"
            if condition == "coherent_block":
                case = corrected_block_case(
                    v1,
                    output,
                    by_scenario_condition[(scenario, "clean")],
                    by_scenario_condition[(scenario, "singleton")],
                    case_id,
                )
            else:
                case = copy_case(
                    v1,
                    output,
                    by_scenario_condition[(scenario, condition)],
                    case_id,
                )
            cases.append(case)

    public_rows = [
        {
            "case_id": case["case_id"],
            "condition": case["condition"],
            "scenario": case["scenario"],
            "workbook": case["workbook"],
        }
        for case in cases
    ]
    with (output / "public" / "manifest.csv").open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(public_rows[0]))
        writer.writeheader()
        writer.writerows(public_rows)

    source_manifest_sha256 = sha256_file(v1 / "SHA256SUMS.txt")
    v2_labels = {
        "protocol": PROTOCOL,
        "derived_from_protocol": labels["protocol"],
        "derived_from_sha256s": source_manifest_sha256,
        "source_archive": labels["source_archive"],
        "source_archive_sha256": labels["source_archive_sha256"],
        "input_scope": "only the five workbooks already present in structuralguard_standard_v1",
        "revealed_development_benchmark": True,
        "conditions": labels["conditions"],
        "cases": cases,
    }
    (output / "labels" / "answer_keys.json").write_text(
        json.dumps(v2_labels, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    metadata = {
        "protocol": PROTOCOL,
        "derived_from_protocol": labels["protocol"],
        "derived_from_sha256s": source_manifest_sha256,
        "source_archive_sha256": labels["source_archive_sha256"],
        "case_count": len(cases),
        "conditions": {
            condition: sum(case["condition"] == condition for case in cases)
            for condition in ("singleton", "clean", "coherent_block", "systematic_column")
        },
        "formula_cells": sum(int(case["formula_cells"]) for case in cases),
        "injected_errors": sum(len(case["errors"]) for case in cases),
        "corrected_v1_coherent_labels": 60,
        "generated_by": "scripts/build_structuralguard_standard_v2.py",
    }
    (output / "metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    lines = []
    for path in sorted(output.rglob("*")):
        if path.is_file() and path.name != "SHA256SUMS.txt":
            lines.append(f"{sha256_file(path)}  {path.relative_to(output).as_posix()}")
    (output / "SHA256SUMS.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(metadata, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
