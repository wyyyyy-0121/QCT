"""Build a fresh multi-family, structurally varied V5.1 confirmation release."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import random
import shutil
import zipfile
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FIXED_TIME = datetime(2000, 1, 1, tzinfo=UTC)
PUBLIC_FIELDS = (
    "case_id",
    "cluster_id",
    "workbook_path",
    "workbook_sha256",
    "file_format",
    "integrity_status",
)
PLAN = (
    ("singleton_a", "singleton", "detect_and_repair"),
    ("singleton_b", "singleton", "detect_and_repair"),
    ("block_a", "contiguous_block", "detect_and_repair"),
    ("block_b", "contiguous_block", "detect_and_repair"),
    ("systematic_a", "systematic_column", "detect_and_repair"),
    ("systematic_b", "systematic_column", "detect_and_repair"),
    ("ambiguous_short", "ambiguous_insufficient", "abstain"),
    ("ambiguous_tied", "ambiguous_tied", "abstain"),
    ("clean_a", "control_clean", "no_action"),
    ("clean_b", "control_clean", "no_action"),
    ("legal_exception", "control_exception", "no_action"),
    ("unsupported_stress", "control_unsupported", "no_action"),
)


def canonical_json(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def anon(seed: str, namespace: str, index: int, length: int) -> str:
    return hashlib.sha256(f"{seed}:{namespace}:{index}".encode()).hexdigest()[:length]


def save_xlsx(workbook: Workbook, path: Path) -> str:
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    raw = BytesIO()
    workbook.save(raw)
    raw.seek(0)
    path.parent.mkdir(parents=True, exist_ok=True)
    with (
        zipfile.ZipFile(raw) as source,
        zipfile.ZipFile(
            path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target,
    ):
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, (1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            target.writestr(info, source.read(name))
    return sha256_file(path)


def style_headers(sheet, row: int, headers: list[str]) -> None:
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(row, column, value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")


def family_workbook(
    rng: random.Random, family: int, cluster: int
) -> tuple[Workbook, dict]:
    workbook = Workbook()
    sheet = workbook.active
    rows = range(6, 28)
    if family == 0:
        sheet.title = "Revenue"
        sheet.merge_cells("A1:G1")
        sheet["A1"] = f"Regional revenue review {cluster + 1}"
        style_headers(
            sheet,
            4,
            ["Period", "Units", "Price", "Revenue", "Cost", "Margin", "Margin %"],
        )
        for row in rows:
            sheet.cell(row, 1, f"2026-{(row - 5):02d}")
            sheet.cell(row, 2, rng.randrange(15, 300))
            sheet.cell(row, 3, round(rng.uniform(4, 90), 2))
            sheet.cell(row, 4, f"=B{row}*C{row}")
            sheet.cell(row, 5, f"=B{row}*C{row}*0.42")
            sheet.cell(row, 6, f"=D{row}-E{row}")
            sheet.cell(row, 7, f"=F{row}/D{row}")
        target = {"sheet": "Revenue", "column": "D", "rows": rows}
    elif family == 1:
        sheet.title = "Inventory"
        sheet["A1"] = "Warehouse stock movement"
        style_headers(
            sheet, 3, ["SKU", "Opening", "Received", "Shipped", "Closing", "Reorder"]
        )
        rows = range(5, 29)
        for row in rows:
            sheet.cell(row, 1, f"SKU-{row:03d}")
            sheet.cell(row, 2, rng.randrange(40, 200))
            sheet.cell(row, 3, rng.randrange(5, 80))
            sheet.cell(row, 4, rng.randrange(2, 70))
            sheet.cell(row, 5, f"=B{row}+C{row}-D{row}")
            sheet.cell(row, 6, f"=IF(E{row}<35,1,0)")
        sheet.row_dimensions[10].hidden = True
        target = {"sheet": "Inventory", "column": "E", "rows": rows}
    elif family == 2:
        sheet.title = "Budget"
        assumptions = workbook.create_sheet("Assumptions")
        assumptions["A1"] = "Overhead rate"
        assumptions["B1"] = 0.18
        sheet.merge_cells("A1:F1")
        sheet["A1"] = "Department budget"
        style_headers(
            sheet, 5, ["Item", "Quantity", "Rate", "Revenue", "Overhead", "Net"]
        )
        rows = range(7, 27)
        for row in rows:
            sheet.cell(row, 1, f"Line {row - 6}")
            sheet.cell(row, 2, rng.randrange(5, 80))
            sheet.cell(row, 3, round(rng.uniform(20, 500), 2))
            sheet.cell(row, 4, f"=B{row}*C{row}")
            sheet.cell(row, 5, f"=D{row}*Assumptions!$B$1")
            sheet.cell(row, 6, f"=D{row}-E{row}")
        sheet["C28"] = "Total"
        for column in range(4, 7):
            letter = chr(64 + column)
            sheet.cell(28, column, f"=SUM({letter}7:{letter}26)")
        target = {"sheet": "Budget", "column": "D", "rows": rows}
    elif family == 3:
        sheet.title = "Grades"
        sheet["A1"] = "Assessment register"
        style_headers(
            sheet, 2, ["Student", "Quiz", "Project", "Exam", "Average", "Pass"]
        )
        rows = range(4, 26)
        for row in rows:
            sheet.cell(row, 1, f"Student {row - 3}")
            for column in range(2, 5):
                sheet.cell(row, column, rng.randrange(45, 100))
            sheet.cell(row, 5, f"=AVERAGE(B{row}:D{row})")
            sheet.cell(row, 6, f"=IF(E{row}>=60,1,0)")
        sheet.merge_cells("H2:I2")
        sheet["H2"] = "Teacher note"
        target = {"sheet": "Grades", "column": "E", "rows": rows}
    elif family == 4:
        sheet.title = "MRR"
        inputs = workbook.create_sheet("Inputs")
        inputs["A1"] = "Default discount"
        inputs["B1"] = 0.05
        sheet["A1"] = "Subscription revenue"
        style_headers(
            sheet,
            6,
            ["Account", "Seats", "Price", "MRR", "Discount", "Net MRR", "Per Seat"],
        )
        rows = range(8, 28)
        for row in rows:
            sheet.cell(row, 1, f"Account-{row - 7:03d}")
            sheet.cell(row, 2, rng.randrange(2, 250))
            sheet.cell(row, 3, round(rng.uniform(8, 120), 2))
            sheet.cell(row, 4, f"=B{row}*C{row}")
            sheet.cell(row, 5, "=Inputs!$B$1")
            sheet.cell(row, 6, f"=D{row}*(1-E{row})")
            sheet.cell(row, 7, f"=F{row}/B{row}")
        target = {"sheet": "MRR", "column": "D", "rows": rows}
    else:
        sheet.title = "Shipments"
        sheet["A1"] = "Dispatch reconciliation"
        style_headers(
            sheet,
            3,
            ["Route", "Opening", "Received", "Delivered", "Closing", "Variance"],
        )
        rows = range(5, 24)
        for row in rows:
            sheet.cell(row, 1, f"Route-{row - 4:02d}")
            sheet.cell(row, 2, rng.randrange(20, 180))
            sheet.cell(row, 3, rng.randrange(5, 90))
            sheet.cell(row, 4, rng.randrange(5, 90))
            sheet.cell(row, 5, f"=B{row}+C{row}-D{row}")
            sheet.cell(row, 6, f"=E{row}-(B{row}+C{row}-D{row})")
        sheet.column_dimensions["A"].width = 18
        target = {"sheet": "Shipments", "column": "E", "rows": rows}
    sheet.freeze_panes = f"A{min(rows) + 1}"
    return workbook, target


def mutate_formula(formula: str) -> str:
    if "*" in formula:
        return formula.replace("*", "+", 1)
    if "-" in formula:
        return formula.replace("-", "+", 1)
    if formula.startswith("=AVERAGE"):
        import re

        match = re.search(r"(\d+)", formula)
        row = match.group(1) if match else "1"
        return f"=B{row}+C{row}"
    return "=0"


def prepare_case(workbook: Workbook, target: dict, kind: str) -> None:
    sheet = workbook[target["sheet"]]
    rows = list(target["rows"])
    if kind == "ambiguous_short":
        for row in rows[2:]:
            for column in range(1, 8):
                sheet.cell(row, column).value = None
    elif kind == "ambiguous_tied":
        for index, row in enumerate(rows):
            formula = sheet[f"{target['column']}{row}"].value
            if index % 2 and isinstance(formula, str):
                sheet[f"{target['column']}{row}"] = mutate_formula(formula)
    elif kind == "legal_exception":
        row = rows[len(rows) // 2]
        formula = sheet[f"{target['column']}{row}"].value
        if isinstance(formula, str) and "*" in formula:
            sheet[f"{target['column']}{row}"] = formula + "+0"
    elif kind == "unsupported_stress":
        row = rows[len(rows) // 2]
        sheet[f"{target['column']}{row}"] = f"=IF(B{row}>0,B{row},0)"


def mutation_rows(target: dict, kind: str) -> list[int]:
    rows = list(target["rows"])
    if kind.startswith("singleton"):
        return [rows[len(rows) // 2]]
    if kind.startswith("block"):
        start = max(1, len(rows) // 3)
        return rows[start : start + min(5, len(rows) - start)]
    if kind.startswith("systematic"):
        return [
            row for index, row in enumerate(rows) if index not in {0, len(rows) - 1}
        ]
    if kind == "ambiguous_short" or kind == "ambiguous_tied":
        return [rows[len(rows) // 2]]
    return []


def deterministic_zip(source: Path, destination: Path, prefix: str) -> None:
    with zipfile.ZipFile(
        destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as archive:
        for path in sorted(item for item in source.rglob("*") if item.is_file()):
            info = zipfile.ZipInfo(
                f"{prefix}/{path.relative_to(source).as_posix()}",
                (1980, 1, 1, 0, 0, 0),
            )
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            archive.writestr(info, path.read_bytes())


def write_sums(root: Path) -> None:
    rows = [
        f"{sha256_file(path)}  {path.relative_to(root).as_posix()}"
        for path in sorted(item for item in root.rglob("*") if item.is_file())
        if path.name != "SHA256SUMS.txt"
    ]
    (root / "SHA256SUMS.txt").write_text("\n".join(rows) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--seed", required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    output = args.output.resolve()
    if output.exists():
        raise SystemExit(f"refusing to overwrite existing release: {output}")
    public = output / "PUBLIC"
    secret = output / "SECRET"
    (public / "workbooks").mkdir(parents=True)
    (secret / "originals").mkdir(parents=True)
    labels: list[dict] = []
    mutation_log: list[dict] = []
    public_rows: list[dict[str, str]] = []
    for cluster_index in range(20):
        family = cluster_index % 6
        cluster_id = "c_" + anon(args.seed, "cluster", cluster_index, 14)
        for plan_index, (kind, cohort, decision) in enumerate(PLAN):
            global_index = cluster_index * len(PLAN) + plan_index
            case_id = "k_" + anon(args.seed, "case", global_index, 20)
            rng = random.Random(int(anon(args.seed, "values", global_index, 16), 16))
            workbook, target = family_workbook(rng, family, cluster_index)
            prepare_case(workbook, target, kind)
            changed_rows = mutation_rows(target, kind)
            errors: list[dict[str, str]] = []
            if changed_rows:
                original_path = secret / "originals" / f"{case_id}.xlsx"
                original_sha = save_xlsx(workbook, original_path)
                sheet = workbook[target["sheet"]]
                for row in changed_rows:
                    address = f"{target['column']}{row}"
                    expected = str(sheet[address].value)
                    injected = mutate_formula(expected)
                    sheet[address] = injected
                    errors.append(
                        {
                            "sheet": target["sheet"],
                            "cell": address,
                            "expected_formula": expected,
                            "injected_formula": injected,
                        }
                    )
            else:
                original_sha = ""
            workbook_path = public / "workbooks" / f"{case_id}.xlsx"
            workbook_sha = save_xlsx(workbook, workbook_path)
            label = {
                "case_id": case_id,
                "cluster_id": cluster_id,
                "family": family,
                "cohort": cohort,
                "decision": decision,
                "workbook_sha256": workbook_sha,
                "original_sha256": original_sha,
                "errors": errors,
            }
            labels.append(label)
            mutation_log.extend({"case_id": case_id, **error} for error in errors)
            public_rows.append(
                {
                    "case_id": case_id,
                    "cluster_id": cluster_id,
                    "workbook_path": f"workbooks/{case_id}.xlsx",
                    "workbook_sha256": workbook_sha,
                    "file_format": "xlsx",
                    "integrity_status": "generator-locked-natural-structure",
                }
            )
    (secret / "labels.json").write_bytes(
        canonical_json(
            {
                "protocol": "v51_natural_confirmation_secret_v1",
                "seed": args.seed,
                "cases": labels,
            }
        )
    )
    (secret / "mutation_log.json").write_bytes(canonical_json(mutation_log))
    (secret / "NOTICE.md").write_text(
        "Fresh multi-family structural confirmation cohort; labels are for disclosed scoring only.\n",
        encoding="utf-8",
    )
    write_sums(secret)
    secret_archive = output / "V51_NATURAL_CONFIRMATION_SECRET_V1.zip"
    deterministic_zip(secret, secret_archive, "SECRET")
    secret_sha = sha256_file(secret_archive)
    (public / "SECRET_ARCHIVE_SHA256.txt").write_text(
        secret_sha + "\n", encoding="ascii"
    )
    with (public / "manifest.csv").open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=PUBLIC_FIELDS)
        writer.writeheader()
        writer.writerows(public_rows)
    write_sums(public)
    public_archive = output / "V51_NATURAL_CONFIRMATION_PUBLIC_V1.zip"
    deterministic_zip(public, public_archive, "PUBLIC")
    public_sha = sha256_file(public_archive)
    (public / "PUBLIC_ARCHIVE_SHA256.txt").write_text(
        public_sha + "\n", encoding="ascii"
    )
    receipt = {
        "protocol": "v51_natural_confirmation_release_v1",
        "generated_at": datetime.now(UTC).isoformat(),
        "cases": 240,
        "clusters": 20,
        "families": 6,
        "public_archive": public_archive.name,
        "public_sha256": public_sha,
        "secret_archive": secret_archive.name,
        "secret_sha256": secret_sha,
        "secret_commitment_inside_public": True,
        "generator_sha256": sha256_file(Path(__file__).resolve()),
    }
    (output / "release_receipt.json").write_bytes(canonical_json(receipt))
    shutil.rmtree(secret)
    print(json.dumps(receipt, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
