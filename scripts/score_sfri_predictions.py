#!/usr/bin/env python3
"""Score locked SFRI predictions against revealed development labels."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import shutil
import statistics
import subprocess
import sys
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from formulaguard.formula import normalized_formula

SCORER_PROTOCOL = "formulaguard_sfri_revealed_score_v1"
PREDICTION_PROTOCOL = "formulaguard_sfri_predictions_v1"
V4_PROTOCOL = "formulaguard_model_discovery_v4_baseline_run_v1"
REVIEW_BUDGET = 5
SELECTED_COHORTS = (
    "enron",
    "public:info1",
    "public:integer_corpus",
    "public:modified_euses",
)
GROUP_FIELDS = (
    "cohort_instance_id",
    "instance_id",
    "cohort",
    "workbook",
    "workbook_sha256",
    "provenance_group_id",
    "structure_cluster_id",
    "outer_group_id",
)
PREDICTION_FIELDS_READ = (
    "cohort",
    "workbook",
    "workbook_sha256",
    "structure_cluster_id",
)
FORBIDDEN_PREDICTION_KEYS = {
    "case_kind",
    "correct_formula",
    "error_type",
    "source_cell",
    "source_cells",
}
FORBIDDEN_PREFIXES = (
    "data/external/v5_psl/revealed_trial",
    "data/external/v5_psl/custodian",
    "data/external/v5_psl/final_blind",
)
SOURCE_PATHS = (
    "formulaguard/formula.py",
    "scripts/score_sfri_predictions.py",
)

DEFAULT_RUN_A = ROOT / "results/sfri_predictions_run_a"
DEFAULT_RUN_B = ROOT / "results/sfri_predictions_run_b"
DEFAULT_V4 = ROOT / "results/model_discovery_v4_baseline"
DEFAULT_GROUPS = ROOT / "results/core_reset_b_phase0/scoring_groups.csv"
DEFAULT_ENRON_LABELS = ROOT / "data/external/enron/manifest.csv"
DEFAULT_PUBLIC_LABELS = (
    ROOT / "results/v5_psl_pressure_inputs/public_pressure_manifest.csv"
)
DEFAULT_PUBLIC_ROOT = ROOT / "results/v5_psl_pressure_inputs"
DEFAULT_OUTPUT = ROOT / "results/sfri_v1_revealed_score"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def canonical_json(payload: object) -> str:
    return json.dumps(
        payload,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    )


def stable_hash(payload: object) -> str:
    return hashlib.sha256(canonical_json(payload).encode("ascii")).hexdigest()


def git_commit(root: Path = ROOT) -> str:
    return subprocess.check_output(
        ("git", "rev-parse", "HEAD"), cwd=root, text=True
    ).strip()


def _git_source_status(root: Path) -> tuple[str, ...]:
    completed = subprocess.run(
        ("git", "status", "--porcelain", "--", *SOURCE_PATHS),
        cwd=root,
        check=True,
        capture_output=True,
        text=True,
    )
    return tuple(line for line in completed.stdout.splitlines() if line)


def capture_source_state(
    source_root: Path = ROOT,
    *,
    allow_dirty: bool = False,
) -> dict[str, object]:
    source_root = source_root.resolve()
    state = {
        "git_commit": git_commit(source_root),
        "source_sha256": {
            relative: sha256(source_root / relative) for relative in SOURCE_PATHS
        },
        "source_status": list(_git_source_status(source_root)),
    }
    dirty = bool(state["source_status"])
    if dirty and not allow_dirty:
        raise ValueError("formal scoring requires clean tracked scorer sources")
    state["source_tree_dirty"] = dirty
    state["formal_evidence"] = not dirty
    return state


def verify_source_state(
    expected: Mapping[str, object], source_root: Path = ROOT
) -> None:
    observed = capture_source_state(source_root, allow_dirty=True)
    if any(
        observed[key] != expected[key]
        for key in ("git_commit", "source_sha256", "source_status")
    ):
        raise ValueError("scorer source changed during scoring")


def _relative(path: Path, *, root: Path = ROOT) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(root.resolve()).as_posix()
    except ValueError:
        return resolved.as_posix()


def _safe_file(path: Path, *, root: Path = ROOT) -> Path:
    candidate = path.resolve()
    resolved_root = root.resolve()
    if candidate != resolved_root and resolved_root not in candidate.parents:
        raise ValueError(f"scoring input is outside the repository: {path}")
    if not candidate.is_file():
        raise FileNotFoundError(candidate)
    relative = candidate.relative_to(resolved_root).as_posix()
    if any(
        relative == prefix or relative.startswith(prefix + "/")
        for prefix in FORBIDDEN_PREFIXES
    ):
        raise ValueError(f"protected input is forbidden: {relative}")
    return candidate


def _load_json(path: Path, *, root: Path = ROOT) -> dict[str, object]:
    value = json.loads(_safe_file(path, root=root).read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise TypeError(f"JSON object required: {path}")
    return value


def _forbidden_keys(payload: object) -> set[str]:
    found: set[str] = set()
    if isinstance(payload, Mapping):
        for key, value in payload.items():
            if str(key) in FORBIDDEN_PREDICTION_KEYS:
                found.add(str(key))
            found.update(_forbidden_keys(value))
    elif isinstance(payload, (list, tuple)):
        for value in payload:
            found.update(_forbidden_keys(value))
    return found


def combined_hash(paths: Iterable[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths, key=lambda item: item.name):
        digest.update(path.name.encode("utf-8"))
        digest.update(b"\0")
        digest.update(path.read_bytes())
        digest.update(b"\0")
    return digest.hexdigest()


def validate_prediction_run(
    run_dir: Path, *, root: Path = ROOT
) -> dict[str, object]:
    receipt_path = _safe_file(run_dir / "completion_receipt.json", root=root)
    summary_path = _safe_file(run_dir / "scan_summary.json", root=root)
    predictions_path = _safe_file(run_dir / "predictions.jsonl", root=root)
    receipt = _load_json(receipt_path, root=root)
    summary = _load_json(summary_path, root=root)
    if receipt.get("protocol") != PREDICTION_PROTOCOL:
        raise ValueError("SFRI receipt protocol mismatch")
    if receipt.get("complete") is not True or receipt.get("formal_evidence") is not True:
        raise ValueError("SFRI prediction run is not complete formal evidence")
    if receipt.get("label_inputs") != [] or receipt.get("protected_data_inputs") != []:
        raise ValueError("SFRI prediction receipt declares forbidden inputs")
    if summary.get("label_inputs") != [] or summary.get("protected_data_inputs") != []:
        raise ValueError("SFRI prediction summary declares forbidden inputs")
    if summary.get("fields_read_from_scoring_groups") != list(PREDICTION_FIELDS_READ):
        raise ValueError("SFRI prediction read an unexpected scoring-group projection")
    if summary.get("selected_cohorts") != list(SELECTED_COHORTS):
        raise ValueError("SFRI prediction cohorts differ from preregistration")
    if receipt.get("scan_summary_sha256") != sha256(summary_path):
        raise ValueError("SFRI summary hash mismatch")
    if receipt.get("predictions_sha256") != sha256(predictions_path):
        raise ValueError("SFRI merged prediction hash mismatch")

    declared_shards = receipt.get("prediction_shard_sha256")
    if not isinstance(declared_shards, Mapping) or not declared_shards:
        raise TypeError("SFRI shard inventory is missing")
    expected_paths = {
        (run_dir / str(relative)).resolve(): str(digest)
        for relative, digest in declared_shards.items()
    }
    observed_paths = set((run_dir / "shards").glob("*.json"))
    if set(expected_paths) != observed_paths:
        raise ValueError("SFRI shard file set differs from the receipt")
    records: list[dict[str, object]] = []
    seen_hashes: set[str] = set()
    for path in sorted(observed_paths, key=lambda item: item.name):
        safe_path = _safe_file(path, root=root)
        if sha256(safe_path) != expected_paths[safe_path]:
            raise ValueError(f"SFRI shard hash mismatch: {path.name}")
        record = _load_json(safe_path, root=root)
        if record.get("protocol") != PREDICTION_PROTOCOL:
            raise ValueError(f"SFRI shard protocol mismatch: {path.name}")
        if record.get("label_inputs") != [] or record.get("protected_data_inputs") != []:
            raise ValueError(f"SFRI shard declares forbidden inputs: {path.name}")
        forbidden = _forbidden_keys(record)
        if forbidden:
            raise ValueError(f"SFRI shard contains label keys: {sorted(forbidden)}")
        workbook_hash = str(record.get("workbook_sha256", ""))
        if path.name != f"{workbook_hash}.json" or workbook_hash in seen_hashes:
            raise ValueError(f"SFRI shard identity mismatch: {path.name}")
        seen_hashes.add(workbook_hash)
        records.append(record)
    records.sort(key=lambda item: str(item["unit_id"]))
    expected_jsonl = "".join(canonical_json(record) + "\n" for record in records)
    if predictions_path.read_text(encoding="ascii") != expected_jsonl:
        raise ValueError("SFRI merged predictions differ from shards")
    if receipt.get("record_set_sha256") != stable_hash(records):
        raise ValueError("SFRI record-set hash mismatch")
    if receipt.get("shard_inventory_sha256") != stable_hash(declared_shards):
        raise ValueError("SFRI shard-inventory hash mismatch")
    if summary.get("prediction_records") != len(records):
        raise ValueError("SFRI summary record count mismatch")
    return {
        "receipt": receipt,
        "summary": summary,
        "records": records,
        "records_by_hash": {str(row["workbook_sha256"]): row for row in records},
        "files": sorted(
            path.relative_to(run_dir.resolve()).as_posix()
            for path in run_dir.resolve().rglob("*")
            if path.is_file()
        ),
    }


def validate_prediction_pair(
    run_a: Path, run_b: Path, *, root: Path = ROOT
) -> tuple[dict[str, object], dict[str, object]]:
    first = validate_prediction_run(run_a, root=root)
    second = validate_prediction_run(run_b, root=root)
    if first["files"] != second["files"]:
        raise ValueError("SFRI prediction runs have different file sets")
    for relative in first["files"]:
        if (run_a / str(relative)).read_bytes() != (run_b / str(relative)).read_bytes():
            raise ValueError(f"SFRI prediction runs differ: {relative}")
    return first, second


def validate_v4_run(run_dir: Path, *, root: Path = ROOT) -> dict[str, object]:
    metadata_path = _safe_file(run_dir / "metadata.json", root=root)
    complete_path = _safe_file(run_dir / "complete.json", root=root)
    metadata = _load_json(metadata_path, root=root)
    complete = _load_json(complete_path, root=root)
    if metadata.get("protocol") != V4_PROTOCOL or complete.get("protocol") != V4_PROTOCOL:
        raise ValueError("V4 baseline protocol mismatch")
    if complete.get("complete") is not True:
        raise ValueError("V4 baseline is incomplete")
    if metadata.get("label_inputs_to_prediction") != []:
        raise ValueError("V4 metadata declares label inputs")
    if complete.get("label_inputs_to_prediction") != []:
        raise ValueError("V4 receipt declares label inputs")
    if complete.get("metadata_sha256") != sha256(metadata_path):
        raise ValueError("V4 metadata hash mismatch")
    paths = sorted((run_dir / "shards").glob("*.json"), key=lambda item: item.name)
    if len(paths) != complete.get("shard_count"):
        raise ValueError("V4 shard count mismatch")
    if complete.get("combined_shards_sha256") != combined_hash(paths):
        raise ValueError("V4 combined shard hash mismatch")
    by_hash: dict[str, dict[str, object]] = {}
    for path in paths:
        record = _load_json(path, root=root)
        if record.get("protocol") != V4_PROTOCOL or record.get("label_inputs") != []:
            raise ValueError(f"V4 shard contract mismatch: {path.name}")
        forbidden = _forbidden_keys(record)
        if forbidden:
            raise ValueError(f"V4 shard contains label keys: {sorted(forbidden)}")
        unhashed = dict(record)
        recorded_hash = unhashed.pop("audit_sha256", None)
        if recorded_hash != stable_hash(unhashed):
            raise ValueError(f"V4 shard audit hash mismatch: {path.name}")
        ranking = record.get("ranking")
        if not isinstance(ranking, list):
            raise TypeError(f"V4 ranking is malformed: {path.name}")
        cells = [str(item.get("cell", "")) for item in ranking if isinstance(item, Mapping)]
        if len(cells) != int(record.get("formula_count", -1)):
            raise ValueError(f"V4 ranking is incomplete: {path.name}")
        if len(cells) != len(set(cells)) or any("!" not in cell for cell in cells):
            raise ValueError(f"V4 ranking cells are malformed: {path.name}")
        workbook_hash = str(record.get("workbook_sha256", ""))
        if workbook_hash in by_hash:
            raise ValueError("V4 baseline contains a duplicate workbook hash")
        by_hash[workbook_hash] = record
    return {
        "metadata": metadata,
        "complete": complete,
        "records_by_hash": by_hash,
    }


def _read_csv(path: Path, *, root: Path, required: Iterable[str]) -> list[dict[str, str]]:
    safe_path = _safe_file(path, root=root)
    with safe_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = set(reader.fieldnames or ())
        missing = set(required) - fields
        if missing:
            raise ValueError(f"{path} missing fields: {sorted(missing)}")
        rows = list(reader)
    if not rows:
        raise ValueError(f"empty CSV: {path}")
    return rows


def _safe_relative_file(base: Path, relative: str, *, root: Path) -> Path:
    value = Path(relative)
    if (
        not relative
        or "\\" in relative
        or value.is_absolute()
        or any(part in {"", ".", ".."} for part in value.parts)
    ):
        raise ValueError(f"unsafe relative input path: {relative!r}")
    candidate = (base / value).resolve()
    resolved_base = base.resolve()
    if resolved_base not in candidate.parents:
        raise ValueError(f"input path escapes its root: {relative!r}")
    return _safe_file(candidate, root=root)


def parse_cells(value: str) -> tuple[str, ...]:
    cells: list[str] = []
    for raw in value.replace("|", ";").split(";"):
        cell = raw.strip()
        if not cell:
            continue
        if "!" not in cell or not cell.rsplit("!", 1)[1]:
            raise ValueError(f"invalid labeled cell: {cell!r}")
        if cell not in cells:
            cells.append(cell)
    return tuple(cells)


def _formula_from_workbook(path: Path, cell: str) -> str | None:
    sheet, address = cell.rsplit("!", 1)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"labeled sheet is absent from original workbook: {cell}")
        value = workbook[sheet][address.replace("$", "")].value
    finally:
        workbook.close()
    return str(value) if isinstance(value, str) and value.startswith("=") else None


def load_revealed_events(
    groups_path: Path,
    enron_labels_path: Path,
    public_labels_path: Path,
    public_root: Path,
    *,
    prediction_records: Mapping[str, object],
    root: Path = ROOT,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    groups_file = _safe_file(groups_path, root=root)
    with groups_file.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if tuple(reader.fieldnames or ()) != GROUP_FIELDS:
            raise ValueError("scoring-group schema differs from the audited contract")
        groups = [row for row in reader if row["cohort"] in SELECTED_COHORTS]
    if len(groups) != 120:
        raise ValueError(f"unexpected selected scoring-group rows: {len(groups)}")
    identifiers = [row["cohort_instance_id"] for row in groups]
    if len(identifiers) != len(set(identifiers)):
        raise ValueError("scoring groups contain duplicate cohort instance IDs")
    if {row["workbook_sha256"] for row in groups} != set(prediction_records):
        raise ValueError("scoring groups do not match the locked prediction workbooks")

    enron_rows = {
        row["instance_id"]: row
        for row in _read_csv(
            enron_labels_path,
            root=root,
            required=("instance_id", "workbook", "source_cells", "source_cell", "correct_formula", "include"),
        )
        if row["include"] == "1"
    }
    public_rows = {
        row["instance_id"]: row
        for row in _read_csv(
            public_labels_path,
            root=root,
            required=("instance_id", "corpus_id", "workbook", "original_workbook", "case_kind", "source_cells", "include"),
        )
        if row["include"] == "1"
    }
    if len(enron_rows) != 30 or len(public_rows) != 90:
        raise ValueError("revealed label event counts differ from the frozen inputs")

    original_inputs: dict[str, str] = {}
    events: list[dict[str, object]] = []
    for group in groups:
        cohort = group["cohort"]
        instance_id = group["instance_id"]
        if cohort == "enron":
            label = enron_rows.get(instance_id)
            if label is None:
                raise ValueError(f"missing Enron label: {instance_id}")
            case_kind = "error"
            sources = parse_cells(label.get("source_cells") or label.get("source_cell", ""))
            correct = label.get("correct_formula", "") if len(sources) == 1 else ""
            expected_path = "data/external/enron/" + label["workbook"]
        else:
            label = public_rows.get(instance_id)
            if label is None or cohort != "public:" + label["corpus_id"]:
                raise ValueError(f"missing or mismatched public label: {instance_id}")
            case_kind = label["case_kind"]
            if case_kind not in {"error", "control"}:
                raise ValueError(f"unexpected public case kind: {case_kind!r}")
            sources = parse_cells(label["source_cells"])
            expected_path = (
                "results/v5_psl_pressure_inputs/" + label["workbook"]
            )
            correct = ""
            if case_kind == "error" and len(sources) == 1:
                original = _safe_relative_file(
                    public_root,
                    label["original_workbook"],
                    root=root,
                )
                original_inputs[_relative(original, root=root)] = sha256(original)
                correct = _formula_from_workbook(original, sources[0]) or ""
        if group["workbook"] != expected_path:
            raise ValueError(f"label workbook mapping mismatch: {instance_id}")
        observed = _safe_file(root / expected_path, root=root)
        if sha256(observed) != group["workbook_sha256"]:
            raise ValueError(f"revealed workbook hash mismatch: {instance_id}")
        if case_kind == "control" and sources:
            raise ValueError(f"control event unexpectedly has source labels: {instance_id}")
        events.append(
            {
                **group,
                "case_kind": case_kind,
                "source_cells": list(sources),
                "correct_formula": correct,
            }
        )
    return events, {
        "scoring_groups": {
            "path": _relative(groups_file, root=root),
            "sha256": sha256(groups_file),
        },
        "revealed_label_files": [
            {
                "path": _relative(_safe_file(path, root=root), root=root),
                "sha256": sha256(_safe_file(path, root=root)),
            }
            for path in (enron_labels_path, public_labels_path)
        ],
        "correct_formula_original_workbooks": {
            "count": len(original_inputs),
            "inventory_sha256": stable_hash(original_inputs),
            "sha256": dict(sorted(original_inputs.items())),
        },
        "protected_data_inputs": [],
    }


def _candidate(record: Mapping[str, object]) -> tuple[str, str] | None:
    result = record.get("result")
    if not isinstance(result, Mapping):
        raise TypeError("SFRI result is malformed")
    candidate = result.get("deterministic_candidate")
    if candidate is None:
        return None
    if not isinstance(candidate, Mapping):
        raise TypeError("SFRI candidate is malformed")
    certificate = candidate.get("certificate")
    if not isinstance(certificate, Mapping):
        raise TypeError("SFRI candidate certificate is malformed")
    raw_cell = certificate.get("target_formula_cell")
    if (
        not isinstance(raw_cell, (list, tuple))
        or len(raw_cell) != 2
        or any(not isinstance(value, str) or not value for value in raw_cell)
    ):
        raise ValueError("SFRI target cell is malformed")
    return f"{raw_cell[0]}!{raw_cell[1]}", str(candidate["candidate_formula"])


def adapt_v4_ranking(
    ranking: Sequence[str], candidate_cell: str | None
) -> tuple[str, ...]:
    cells = tuple(ranking)
    if len(cells) != len(set(cells)):
        raise ValueError("V4 ranking contains duplicate cells")
    if candidate_cell is None or candidate_cell in cells[:REVIEW_BUDGET]:
        return cells
    if candidate_cell not in cells:
        raise ValueError("SFRI candidate is absent from the V4 ranking")
    if len(cells) <= REVIEW_BUDGET:
        return cells
    return (
        *cells[: REVIEW_BUDGET - 1],
        candidate_cell,
        *(cell for cell in cells[REVIEW_BUDGET - 1 :] if cell != candidate_cell),
    )


def ranking_metric(
    ranking: Sequence[str], source_cells: Sequence[str]
) -> dict[str, int | float | None]:
    positions = {cell: index for index, cell in enumerate(ranking, 1)}
    ranks = [positions[cell] for cell in source_cells if cell in positions]
    rank = min(ranks) if ranks else None
    return {
        "rank": rank,
        "top1": int(rank is not None and rank <= 1),
        "top5": int(rank is not None and rank <= REVIEW_BUDGET),
        "mrr": 1.0 / rank if rank is not None else 0.0,
    }


def _metric_summary(rows: Sequence[Mapping[str, object]], method: str) -> dict[str, object]:
    selected = [row for row in rows if row["case_kind"] == "error"]
    return {
        "events": len(selected),
        **{
            key: statistics.fmean(
                float(row["metrics"][method][key]) for row in selected  # type: ignore[index]
            )
            if selected
            else None
            for key in ("top1", "top5", "mrr")
        },
    }


def _structure_metric_summary(
    rows: Sequence[Mapping[str, object]], method: str
) -> dict[str, object]:
    groups: dict[str, list[Mapping[str, object]]] = defaultdict(list)
    for row in rows:
        if row["case_kind"] == "error":
            groups[str(row["structure_cluster_id"])].append(row)
    return {
        "groups": len(groups),
        "events": sum(len(values) for values in groups.values()),
        **{
            key: statistics.fmean(
                statistics.fmean(
                    float(row["metrics"][method][key])  # type: ignore[index]
                    for row in values
                )
                for values in groups.values()
            )
            if groups
            else None
            for key in ("top1", "top5", "mrr")
        },
    }


def evaluate_gates(
    *,
    candidate_workbooks: int,
    candidate_structure_clusters: int,
    certificate_precision: float | None,
    false_control_workbooks: int,
    formula_accuracy: float | None,
    net_top5_rescues: int,
    newly_hit_structure_clusters: int,
    top5_losses: int,
    cohort_top5_losses: Mapping[str, int],
    enron_top5_delta: float,
    enron_mrr_delta: float,
    public_control_action_rate: float,
    integrity_passed: bool,
) -> dict[str, object]:
    gates = {
        "g1_coverage": candidate_workbooks >= 6
        and candidate_structure_clusters >= 2,
        "g2_certificate_precision_and_controls": (
            certificate_precision is not None
            and certificate_precision >= 0.90
            and false_control_workbooks == 0
        ),
        "g3_candidate_formula_accuracy": (
            formula_accuracy is not None and formula_accuracy >= 0.90
        ),
        "g4_net_top5_rescues": (
            net_top5_rescues >= 2 and newly_hit_structure_clusters >= 1
        ),
        "g5_zero_top5_losses": top5_losses == 0
        and all(value == 0 for value in cohort_top5_losses.values()),
        "g6_enron_and_control_safety": (
            enron_top5_delta >= 0
            and enron_mrr_delta >= 0
            and public_control_action_rate == 0
        ),
        "g7_integrity": integrity_passed,
    }
    return {
        **gates,
        "all_gates_passed": all(gates.values()),
        "failed_gates": sorted(key for key, value in gates.items() if not value),
    }


def build_scores(
    events: Sequence[Mapping[str, object]],
    prediction_by_hash: Mapping[str, object],
    v4_by_hash: Mapping[str, object],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    rows: list[dict[str, object]] = []
    for event in events:
        workbook_hash = str(event["workbook_sha256"])
        prediction = prediction_by_hash.get(workbook_hash)
        baseline = v4_by_hash.get(workbook_hash)
        if not isinstance(prediction, Mapping) or not isinstance(baseline, Mapping):
            raise TypeError("an event lacks a locked SFRI or V4 prediction")
        ranking_rows = baseline.get("ranking")
        if not isinstance(ranking_rows, list):
            raise TypeError("V4 ranking is malformed")
        v4_ranking = tuple(str(item["cell"]) for item in ranking_rows)
        candidate = _candidate(prediction)
        candidate_cell = candidate[0] if candidate else None
        candidate_formula = candidate[1] if candidate else None
        adapted = adapt_v4_ranking(v4_ranking, candidate_cell)
        sources = tuple(str(cell) for cell in event["source_cells"])
        case_kind = str(event["case_kind"])
        if case_kind == "error":
            metrics = {
                "v4_r1": ranking_metric(v4_ranking, sources),
                "v4_sfri_fifth": ranking_metric(adapted, sources),
            }
        else:
            metrics = {"v4_r1": None, "v4_sfri_fifth": None}
        correct = str(event.get("correct_formula", ""))
        formula_evaluable = bool(
            candidate_formula and candidate_cell in sources and correct
        )
        formula_exact = (
            normalized_formula(candidate_formula) == normalized_formula(correct)
            if formula_evaluable and candidate_formula is not None
            else None
        )
        rows.append(
            {
                "cohort_instance_id": event["cohort_instance_id"],
                "instance_id": event["instance_id"],
                "cohort": event["cohort"],
                "case_kind": case_kind,
                "workbook": event["workbook"],
                "workbook_sha256": workbook_hash,
                "structure_cluster_id": event["structure_cluster_id"],
                "source_cells": list(sources),
                "candidate_cell": candidate_cell,
                "candidate_formula": candidate_formula,
                "certificate_emitted": candidate is not None,
                "certificate_target_hit": (
                    int(candidate_cell in sources)
                    if candidate_cell is not None
                    else None
                ),
                "correct_formula_available": formula_evaluable,
                "candidate_formula_exact": formula_exact,
                "metrics": metrics,
            }
        )
    rows.sort(key=lambda item: str(item["cohort_instance_id"]))

    error_rows = [row for row in rows if row["case_kind"] == "error"]
    control_rows = [row for row in rows if row["case_kind"] == "control"]
    rescues = [
        row
        for row in error_rows
        if row["metrics"]["v4_r1"]["top5"] == 0  # type: ignore[index]
        and row["metrics"]["v4_sfri_fifth"]["top5"] == 1  # type: ignore[index]
    ]
    losses = [
        row
        for row in error_rows
        if row["metrics"]["v4_r1"]["top5"] == 1  # type: ignore[index]
        and row["metrics"]["v4_sfri_fifth"]["top5"] == 0  # type: ignore[index]
    ]
    by_cohort_losses = {
        cohort: sum(row["cohort"] == cohort for row in losses)
        for cohort in SELECTED_COHORTS
    }

    events_by_hash: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        events_by_hash[str(row["workbook_sha256"])].append(row)
    candidate_records = {
        workbook_hash: record
        for workbook_hash, record in prediction_by_hash.items()
        if isinstance(record, Mapping) and _candidate(record) is not None
    }
    candidate_workbook_rows: list[dict[str, object]] = []
    for workbook_hash, prediction in candidate_records.items():
        if not isinstance(prediction, Mapping):
            raise TypeError("SFRI prediction is malformed")
        candidate = _candidate(prediction)
        assert candidate is not None
        candidate_cell, candidate_formula = candidate
        related = events_by_hash[workbook_hash]
        kinds = {str(row["case_kind"]) for row in related}
        if len(kinds) != 1:
            raise ValueError("one observed workbook maps to mixed error/control events")
        hit_rows = [
            row
            for row in related
            if candidate_cell in row["source_cells"]
        ]
        evaluable = [row for row in hit_rows if row["correct_formula_available"]]
        candidate_workbook_rows.append(
            {
                "workbook_sha256": workbook_hash,
                "cohort": prediction["cohort"],
                "structure_cluster_id": prediction["structure_cluster_id"],
                "case_kind": next(iter(kinds)),
                "candidate_cell": candidate_cell,
                "candidate_formula": candidate_formula,
                "target_hit": bool(hit_rows),
                "formula_evaluable": bool(evaluable),
                "candidate_formula_exact": (
                    all(row["candidate_formula_exact"] is True for row in evaluable)
                    if evaluable
                    else None
                ),
                "related_event_ids": sorted(str(row["instance_id"]) for row in related),
            }
        )
    candidate_workbook_rows.sort(key=lambda item: str(item["workbook_sha256"]))

    precision_numerator = sum(
        row["target_hit"] is True for row in candidate_workbook_rows
    )
    precision_denominator = len(candidate_workbook_rows)
    formula_rows = [
        row for row in candidate_workbook_rows if row["formula_evaluable"] is True
    ]
    control_candidate_rows = [
        row for row in candidate_workbook_rows if row["case_kind"] == "control"
    ]
    public_controls = [
        row for row in control_rows if str(row["cohort"]).startswith("public:")
    ]
    enron_rows = [row for row in rows if row["cohort"] == "enron"]
    overall_metrics = {
        method: {
            "event_macro": _metric_summary(rows, method),
            "structure_group_macro": _structure_metric_summary(rows, method),
        }
        for method in ("v4_r1", "v4_sfri_fifth")
    }
    cohort_metrics = {
        cohort: {
            method: {
                "event_macro": _metric_summary(
                    [row for row in rows if row["cohort"] == cohort], method
                ),
                "structure_group_macro": _structure_metric_summary(
                    [row for row in rows if row["cohort"] == cohort], method
                ),
            }
            for method in ("v4_r1", "v4_sfri_fifth")
        }
        for cohort in SELECTED_COHORTS
    }
    certificate_precision = (
        precision_numerator / precision_denominator if precision_denominator else None
    )
    formula_accuracy = (
        sum(row["candidate_formula_exact"] is True for row in formula_rows)
        / len(formula_rows)
        if formula_rows
        else None
    )
    public_control_action_rate = (
        sum(row["certificate_emitted"] is True for row in public_controls)
        / len(public_controls)
        if public_controls
        else 0.0
    )
    enron_v4 = _metric_summary(enron_rows, "v4_r1")
    enron_sfri = _metric_summary(enron_rows, "v4_sfri_fifth")
    paired = {
        "top5_rescues": len(rescues),
        "top5_losses": len(losses),
        "net_top5_rescues": len(rescues) - len(losses),
        "rescue_event_ids": [str(row["instance_id"]) for row in rescues],
        "loss_event_ids": [str(row["instance_id"]) for row in losses],
        "newly_hit_structure_clusters": sorted(
            {str(row["structure_cluster_id"]) for row in rescues}
        ),
        "top5_losses_by_cohort": by_cohort_losses,
    }
    certificate_summary = {
        "candidate_workbooks": len(candidate_workbook_rows),
        "candidate_structure_clusters": len(
            {str(row["structure_cluster_id"]) for row in candidate_workbook_rows}
        ),
        "target_hits": precision_numerator,
        "target_precision": certificate_precision,
        "false_control_workbooks": len(control_candidate_rows),
        "formula_evaluable_workbooks": len(formula_rows),
        "candidate_formula_exact_workbooks": sum(
            row["candidate_formula_exact"] is True for row in formula_rows
        ),
        "candidate_formula_accuracy": formula_accuracy,
        "workbooks": candidate_workbook_rows,
    }
    control_summary = {
        "public_control_events": len(public_controls),
        "public_control_action_events": sum(
            row["certificate_emitted"] is True for row in public_controls
        ),
        "public_control_action_rate": public_control_action_rate,
        "public_control_workbooks": len(
            {str(row["workbook_sha256"]) for row in public_controls}
        ),
        "public_control_action_workbooks": len(control_candidate_rows),
    }
    gates = evaluate_gates(
        candidate_workbooks=len(candidate_workbook_rows),
        candidate_structure_clusters=certificate_summary[
            "candidate_structure_clusters"
        ],
        certificate_precision=certificate_precision,
        false_control_workbooks=len(control_candidate_rows),
        formula_accuracy=formula_accuracy,
        net_top5_rescues=paired["net_top5_rescues"],
        newly_hit_structure_clusters=len(paired["newly_hit_structure_clusters"]),
        top5_losses=len(losses),
        cohort_top5_losses=by_cohort_losses,
        enron_top5_delta=float(enron_sfri["top5"]) - float(enron_v4["top5"]),
        enron_mrr_delta=float(enron_sfri["mrr"]) - float(enron_v4["mrr"]),
        public_control_action_rate=public_control_action_rate,
        integrity_passed=True,
    )
    return rows, {
        "event_count": len(rows),
        "error_events": len(error_rows),
        "control_events": len(control_rows),
        "overall_metrics": overall_metrics,
        "cohort_metrics": cohort_metrics,
        "certificate_summary": certificate_summary,
        "paired_top5": paired,
        "control_summary": control_summary,
        "enron_safety": {
            "v4_top5": enron_v4["top5"],
            "sfri_top5": enron_sfri["top5"],
            "top5_delta": float(enron_sfri["top5"]) - float(enron_v4["top5"]),
            "v4_mrr": enron_v4["mrr"],
            "sfri_mrr": enron_sfri["mrr"],
            "mrr_delta": float(enron_sfri["mrr"]) - float(enron_v4["mrr"]),
        },
        "gates": gates,
    }


def write_score(
    output: Path,
    rows: Sequence[Mapping[str, object]],
    summary: Mapping[str, object],
) -> Path:
    output = output.resolve()
    partial = output.with_name(output.name + ".partial")
    if output.exists() or partial.exists():
        raise ValueError("score output or partial output already exists")
    partial.mkdir(parents=True)
    try:
        events_path = partial / "event_scores.jsonl"
        with events_path.open("w", encoding="ascii", newline="\n") as handle:
            for row in rows:
                handle.write(canonical_json(row) + "\n")
        payload = {
            **dict(summary),
            "event_scores_sha256": sha256(events_path),
            "event_score_set_sha256": stable_hash(rows),
        }
        summary_path = partial / "score_summary.json"
        summary_path.write_text(
            json.dumps(payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n",
            encoding="ascii",
        )
        os.replace(partial, output)
    except BaseException:
        shutil.rmtree(partial, ignore_errors=True)
        raise
    return output / "score_summary.json"


def score(
    *,
    run_a: Path,
    run_b: Path,
    v4_dir: Path,
    groups: Path,
    enron_labels: Path,
    public_labels: Path,
    public_root: Path,
    output: Path,
    root: Path = ROOT,
    source_root: Path = ROOT,
    allow_dirty: bool = False,
) -> Path:
    source_state = capture_source_state(source_root, allow_dirty=allow_dirty)

    # Prediction and baseline integrity must pass before revealed labels are read.
    first, second = validate_prediction_pair(run_a, run_b, root=root)
    baseline = validate_v4_run(v4_dir, root=root)
    prediction_by_hash = first["records_by_hash"]
    v4_by_hash = baseline["records_by_hash"]
    if not isinstance(prediction_by_hash, Mapping) or not isinstance(v4_by_hash, Mapping):
        raise TypeError("locked prediction indexes are malformed")

    events, label_audit = load_revealed_events(
        groups,
        enron_labels,
        public_labels,
        public_root,
        prediction_records=prediction_by_hash,
        root=root,
    )
    rows, metrics = build_scores(events, prediction_by_hash, v4_by_hash)
    verify_source_state(source_state, source_root)
    first_receipt = first["receipt"]
    second_receipt = second["receipt"]
    if not isinstance(first_receipt, Mapping) or not isinstance(second_receipt, Mapping):
        raise TypeError("prediction receipts are malformed")
    complete = baseline["complete"]
    if not isinstance(complete, Mapping):
        raise TypeError("V4 completion receipt is malformed")
    summary = {
        "protocol": SCORER_PROTOCOL,
        "complete": True,
        **dict(source_state),
        "prediction_integrity_verified_before_labels": True,
        "prediction_runs_byte_identical": True,
        "prediction_runs": [
            {
                "path": _relative(run, root=root),
                "completion_receipt_sha256": sha256(
                    run / "completion_receipt.json"
                ),
                "git_commit": receipt["git_commit"],
                "record_set_sha256": receipt["record_set_sha256"],
                "shard_inventory_sha256": receipt["shard_inventory_sha256"],
            }
            for run, receipt in (
                (run_a, first_receipt),
                (run_b, second_receipt),
            )
        ],
        "v4_baseline": {
            "path": _relative(v4_dir, root=root),
            "metadata_sha256": sha256(v4_dir / "metadata.json"),
            "complete_sha256": sha256(v4_dir / "complete.json"),
            "combined_shards_sha256": complete["combined_shards_sha256"],
        },
        "label_audit": label_audit,
        "fixed_rule": "preserve_v4_top4_then_place_one_sfri_candidate_fifth",
        "review_budget": REVIEW_BUDGET,
        **metrics,
        "decision": {
            "bounded_candidate_authorized": metrics["gates"]["all_gates_passed"],
            "formal_v5_r1_authorized": False,
            "negative_result_required": not metrics["gates"]["all_gates_passed"],
            "post_score_rule_changes_forbidden": True,
        },
        "protected_data_inputs": [],
    }
    return write_score(output, rows, summary)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run-a", type=Path, default=DEFAULT_RUN_A)
    parser.add_argument("--run-b", type=Path, default=DEFAULT_RUN_B)
    parser.add_argument("--v4-dir", type=Path, default=DEFAULT_V4)
    parser.add_argument("--groups", type=Path, default=DEFAULT_GROUPS)
    parser.add_argument("--enron-labels", type=Path, default=DEFAULT_ENRON_LABELS)
    parser.add_argument("--public-labels", type=Path, default=DEFAULT_PUBLIC_LABELS)
    parser.add_argument("--public-root", type=Path, default=DEFAULT_PUBLIC_ROOT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--allow-dirty", action="store_true")
    args = parser.parse_args(argv)
    try:
        result = score(
            run_a=args.run_a,
            run_b=args.run_b,
            v4_dir=args.v4_dir,
            groups=args.groups,
            enron_labels=args.enron_labels,
            public_labels=args.public_labels,
            public_root=args.public_root,
            output=args.output,
            allow_dirty=args.allow_dirty,
        )
    except (OSError, TypeError, ValueError, KeyError, json.JSONDecodeError) as exc:
        raise SystemExit(f"SFRI scoring refused: {exc}") from exc
    print(f"SFRI score: {result}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
