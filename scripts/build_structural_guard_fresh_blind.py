"""Generate one fresh 360-case PUBLIC/SECRET Structural Guard blind release."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import random
import shutil
import zipfile
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FIXED_TIME = datetime(2000, 1, 1, tzinfo=UTC)
PUBLIC_FIELDS = (
    "case_id",
    "cluster_id",
    "workbook_path",
    "workbook_sha256",
    "file_format",
    "integrity_status",
)
LANGUAGES = (
    (
        "en",
        "Operations",
        ("Period", "Units", "Price", "Revenue", "Cost", "Margin", "Average"),
    ),
    ("zh", "运营", ("期间", "数量", "单价", "收入", "成本", "利润", "平均值")),
    (
        "es",
        "Operaciones",
        ("Periodo", "Unidades", "Precio", "Ingresos", "Coste", "Margen", "Promedio"),
    ),
)
CASE_PLAN = (
    ("singleton_revenue", "singleton", "detect_and_repair"),
    ("singleton_margin", "singleton", "detect_and_repair"),
    ("block_revenue", "contiguous_block", "detect_and_repair"),
    ("block_margin", "contiguous_block", "detect_and_repair"),
    ("systematic_revenue", "systematic_column", "detect_and_repair"),
    ("systematic_margin", "systematic_column", "detect_and_repair"),
    ("ambiguous_short", "ambiguous_insufficient", "abstain"),
    ("ambiguous_tied", "ambiguous_tied", "abstain"),
    ("clean_regular_a", "control_clean", "no_action"),
    ("clean_regular_b", "control_clean", "no_action"),
    ("legal_exception", "control_exception", "no_action"),
    ("unsupported_stress", "control_unsupported", "no_action"),
)


def canonical_json(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def token(seed: str, namespace: str, index: int, length: int) -> str:
    return hashlib.sha256(f"{seed}:{namespace}:{index}".encode()).hexdigest()[:length]


def canonical_xlsx(wb: Workbook, path: Path) -> str:
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    raw = BytesIO()
    wb.save(raw)
    raw.seek(0)
    path.parent.mkdir(parents=True, exist_ok=True)
    with (
        zipfile.ZipFile(raw) as source,
        zipfile.ZipFile(
            path,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=9,
        ) as target,
    ):
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, (1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            target.writestr(info, source.read(name))
    return sha256_file(path)


def make_clean_workbook(
    rng: random.Random, language_index: int, rows: int = 34
) -> Workbook:
    _, title, headers = LANGUAGES[language_index]
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws["A1"] = f"{title} {rng.randrange(1000, 9999)}"
    ws["A2"] = "Synthetic blind evaluation workbook"
    for column, header in enumerate(headers, 1):
        cell = ws.cell(4, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws["I4"] = "Tax"
    ws["I5"] = round(rng.uniform(0.04, 0.12), 3)
    for row in range(5, rows + 5):
        ws.cell(row, 1, f"P{row - 4:02d}")
        ws.cell(row, 2, rng.randrange(25, 500))
        ws.cell(row, 3, round(rng.uniform(3.5, 95.0), 2))
        ws.cell(row, 4, f"=B{row}*C{row}")
        ws.cell(row, 5, f"=B{row}*C{row}*(0.45+$I$5)")
        ws.cell(row, 6, f"=D{row}-E{row}")
        ws.cell(row, 7, f"=AVERAGE(D{row}:F{row})")
    total = rows + 6
    ws.cell(total, 3, "Total")
    for column in range(4, 8):
        letter = get_column_letter(column)
        ws.cell(total, column, f"=SUM({letter}5:{letter}{rows + 4})")
    for column, width in {1: 13, 2: 12, 3: 12, 4: 15, 5: 15, 6: 15, 7: 15}.items():
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A5"
    return wb


def prepare_case(wb: Workbook, kind: str) -> None:
    ws = wb.active
    if kind == "ambiguous_short":
        for row in range(9, 39):
            for column in range(1, 8):
                ws.cell(row, column).value = None
    elif kind == "ambiguous_tied":
        for row in range(5, 39):
            ws[f"D{row}"] = f"=B{row}*C{row}" if row % 2 else f"=B{row}*C{row}+$I$5"
    elif kind == "clean_regular_b":
        for row in range(5, 39):
            ws[f"G{row}"] = f"=SUM(D{row}:F{row})/3"
    elif kind == "legal_exception":
        ws["D21"] = "=B21*C21+$I$5"
        ws["A21"] = "P17 adjustment"
    elif kind == "unsupported_stress":
        for row in range(5, 39):
            ws[f"G{row}"] = f"=IF(B{row}>0,D{row}-E{row},0)"


def apply_mutation(wb: Workbook, kind: str) -> list[dict[str, str]]:
    ws = wb.active
    changes: list[tuple[str, str]] = []
    if kind == "singleton_revenue":
        changes = [("D17", "=B17+C17")]
    elif kind == "singleton_margin":
        changes = [("F24", "=D24+E24")]
    elif kind == "block_revenue":
        changes = [(f"D{row}", f"=B{row}+C{row}") for row in range(14, 20)]
    elif kind == "block_margin":
        changes = [(f"F{row}", f"=D{row}+E{row}") for row in range(22, 29)]
    elif kind == "systematic_revenue":
        changes = [(f"D{row}", f"=B{row}+C{row}") for row in range(8, 36)]
    elif kind == "systematic_margin":
        changes = [(f"F{row}", f"=D{row}+E{row}") for row in range(7, 37)]
    elif kind == "ambiguous_short":
        changes = [("F6", "=D6+E6")]
    elif kind == "ambiguous_tied":
        changes = [("D21", "=B21+C21")]
    records = []
    for cell, injected in changes:
        expected = str(ws[cell].value)
        ws[cell] = injected
        records.append(
            {
                "sheet": ws.title,
                "cell": cell,
                "expected_formula": expected,
                "injected_formula": injected,
            }
        )
    return records


def deterministic_zip(source: Path, destination: Path, prefix: str) -> None:
    with zipfile.ZipFile(
        destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as bundle:
        for path in sorted(item for item in source.rglob("*") if item.is_file()):
            info = zipfile.ZipInfo(
                f"{prefix}/{path.relative_to(source).as_posix()}", (1980, 1, 1, 0, 0, 0)
            )
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            bundle.writestr(info, path.read_bytes())


def write_sums(root: Path, exclude: set[str] | None = None) -> None:
    excluded = exclude or set()
    rows = []
    for path in sorted(item for item in root.rglob("*") if item.is_file()):
        relative = path.relative_to(root).as_posix()
        if relative not in excluded:
            rows.append(f"{sha256_file(path)}  {relative}")
    (root / "SHA256SUMS.txt").write_text("\n".join(rows) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--seed",
        required=True,
        help="single-use random seed chosen after protocol lock",
    )
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    output = args.output.resolve()
    if output.exists():
        raise SystemExit(f"refusing to overwrite existing release: {output}")
    public = output / "PUBLIC"
    secret = output / "SECRET"
    originals = secret / "originals"
    public_workbooks = public / "workbooks"
    originals.mkdir(parents=True)
    public_workbooks.mkdir(parents=True)
    labels = []
    mutation_log = []
    public_rows = []
    for cluster_index in range(30):
        language_index = cluster_index // 10
        cluster_id = "c_" + token(args.seed, "cluster", cluster_index, 14)
        for plan_index, (kind, cohort, decision) in enumerate(CASE_PLAN):
            global_index = cluster_index * len(CASE_PLAN) + plan_index
            case_id = "k_" + token(args.seed, "case", global_index, 20)
            case_rng = random.Random(
                int(token(args.seed, "values", global_index, 16), 16)
            )
            clean = make_clean_workbook(case_rng, language_index)
            prepare_case(clean, kind)
            errors_expected = kind in {
                "singleton_revenue",
                "singleton_margin",
                "block_revenue",
                "block_margin",
                "systematic_revenue",
                "systematic_margin",
                "ambiguous_short",
                "ambiguous_tied",
            }
            if errors_expected:
                original_path = originals / f"{case_id}.xlsx"
                original_sha = canonical_xlsx(clean, original_path)
            else:
                original_sha = ""
            changes = apply_mutation(clean, kind)
            if bool(changes) != errors_expected:
                raise ValueError(f"case mutation mismatch: {kind}")
            workbook_path = public_workbooks / f"{case_id}.xlsx"
            workbook_sha = canonical_xlsx(clean, workbook_path)
            label = {
                "case_id": case_id,
                "cluster_id": cluster_id,
                "language": LANGUAGES[language_index][0],
                "cohort": cohort,
                "decision": decision,
                "workbook_sha256": workbook_sha,
                "original_sha256": original_sha,
                "errors": changes,
            }
            labels.append(label)
            for change in changes:
                mutation_log.append({"case_id": case_id, **change})
            public_rows.append(
                {
                    "case_id": case_id,
                    "cluster_id": cluster_id,
                    "workbook_path": f"workbooks/{case_id}.xlsx",
                    "workbook_sha256": workbook_sha,
                    "file_format": "xlsx",
                    "integrity_status": "generator-locked-canonical-xlsx",
                }
            )
    (secret / "labels.json").write_bytes(
        canonical_json(
            {
                "protocol": "structural_guard_fresh_blind_secret_v1",
                "seed": args.seed,
                "cases": labels,
            }
        )
    )
    (secret / "mutation_log.json").write_bytes(canonical_json(mutation_log))
    (secret / "NOTICE.md").write_text(
        "Fresh AI-administered synthetic blind cohort. Models and metrics were frozen before seed selection.\n",
        encoding="utf-8",
    )
    write_sums(secret)
    secret_archive = output / "STRUCTURAL_GUARD_FRESH_SECRET_V1.zip"
    deterministic_zip(secret, secret_archive, "SECRET")
    secret_sha = sha256_file(secret_archive)
    (public / "SECRET_ARCHIVE_SHA256.txt").write_text(
        secret_sha + "\n", encoding="ascii"
    )
    with (public / "manifest.csv").open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=PUBLIC_FIELDS)
        writer.writeheader()
        writer.writerows(public_rows)
    write_sums(public)
    # PUBLIC includes this receipt only after its archive is complete; the runner
    # verifies the copied self-hash receipt supplied beside the extracted tree.
    public_archive = output / "STRUCTURAL_GUARD_FRESH_PUBLIC_V1.zip"
    deterministic_zip(public, public_archive, "PUBLIC")
    public_sha = sha256_file(public_archive)
    (public / "PUBLIC_ARCHIVE_SHA256.txt").write_text(
        public_sha + "\n", encoding="ascii"
    )
    receipt = {
        "protocol": "structural_guard_fresh_blind_release_v1",
        "generated_at": datetime.now(UTC).isoformat(),
        "cases": 360,
        "clusters": 30,
        "public_archive": public_archive.name,
        "public_sha256": public_sha,
        "secret_archive": secret_archive.name,
        "secret_sha256": secret_sha,
        "secret_commitment_inside_public": True,
        "generator_sha256": sha256_file(Path(__file__).resolve()),
    }
    (output / "release_receipt.json").write_bytes(canonical_json(receipt))
    # Remove extracted SECRET to enforce the reveal boundary. The committed ZIP
    # remains available to the scorer only after prediction locks exist.
    shutil.rmtree(secret)
    print(json.dumps(receipt, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
