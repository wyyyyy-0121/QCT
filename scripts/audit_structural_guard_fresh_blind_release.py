"""Audit PUBLIC/SECRET identity and declared workbook mutations after reveal."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def workbook_values(
    payload: bytes,
) -> tuple[tuple[str, ...], dict[tuple[str, str], object]]:
    workbook = load_workbook(BytesIO(payload), data_only=False)
    values = {
        (sheet.title, cell.coordinate): cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    return tuple(workbook.sheetnames), values


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--release", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    receipt = json.loads((args.release / "release_receipt.json").read_text())
    public_archive = args.release / receipt["public_archive"]
    secret_archive = args.release / receipt["secret_archive"]
    if sha256_file(public_archive) != receipt["public_sha256"]:
        raise SystemExit("PUBLIC archive hash failure")
    if sha256_file(secret_archive) != receipt["secret_sha256"]:
        raise SystemExit("SECRET archive hash failure")
    with zipfile.ZipFile(secret_archive) as secret:
        labels_payload = secret.read("SECRET/labels.json")
        labels = json.loads(labels_payload)["cases"]
        mutation_log = json.loads(secret.read("SECRET/mutation_log.json"))
        secret_names = set(secret.namelist())
        original_payloads = {
            case["case_id"]: secret.read(f"SECRET/originals/{case['case_id']}.xlsx")
            for case in labels
            if case["errors"]
        }
    with (args.release / "PUBLIC" / "manifest.csv").open(
        encoding="utf-8", newline=""
    ) as stream:
        public_rows = list(csv.DictReader(stream))
    labels_by_id = {case["case_id"]: case for case in labels}
    if set(labels_by_id) != {row["case_id"] for row in public_rows}:
        raise SystemExit("PUBLIC/SECRET case identity failure")
    undeclared_formula_differences = 0
    nonformula_differences = 0
    sheet_structure_differences = 0
    checked_error_cases = 0
    checked_changed_cells = 0
    for row in public_rows:
        case = labels_by_id[row["case_id"]]
        public_path = args.release / "PUBLIC" / row["workbook_path"]
        if sha256_file(public_path) != row["workbook_sha256"]:
            raise SystemExit(f"PUBLIC workbook hash failure: {row['case_id']}")
        if not case["errors"]:
            continue
        original = original_payloads[row["case_id"]]
        if sha256_bytes(original) != case["original_sha256"]:
            raise SystemExit(f"original hash failure: {row['case_id']}")
        original_sheets, original_values = workbook_values(original)
        public_sheets, public_values = workbook_values(public_path.read_bytes())
        if original_sheets != public_sheets:
            sheet_structure_differences += 1
        declared = {(item["sheet"], item["cell"]): item for item in case["errors"]}
        differences = {
            key
            for key in set(original_values) | set(public_values)
            if original_values.get(key) != public_values.get(key)
        }
        if differences != set(declared):
            undeclared_formula_differences += len(differences - set(declared))
            nonformula_differences += sum(
                not str(original_values.get(key, "")).startswith("=")
                or not str(public_values.get(key, "")).startswith("=")
                for key in differences - set(declared)
            )
        for key, item in declared.items():
            if original_values.get(key) != item["expected_formula"]:
                raise SystemExit(f"expected formula mismatch: {row['case_id']} {key}")
            if public_values.get(key) != item["injected_formula"]:
                raise SystemExit(f"injected formula mismatch: {row['case_id']} {key}")
        checked_error_cases += 1
        checked_changed_cells += len(declared)
    ledger = {
        (
            item["case_id"],
            item["sheet"],
            item["cell"],
            item["expected_formula"],
            item["injected_formula"],
        )
        for item in mutation_log
    }
    declared_ledger = {
        (
            case["case_id"],
            item["sheet"],
            item["cell"],
            item["expected_formula"],
            item["injected_formula"],
        )
        for case in labels
        for item in case["errors"]
    }
    result = {
        "protocol": "structural_guard_fresh_blind_post_reveal_audit_v1",
        "passed": (
            ledger == declared_ledger
            and undeclared_formula_differences == 0
            and nonformula_differences == 0
            and sheet_structure_differences == 0
        ),
        "public_sha256": receipt["public_sha256"],
        "secret_sha256": receipt["secret_sha256"],
        "cases": len(labels),
        "clusters": len({case["cluster_id"] for case in labels}),
        "error_cases_checked": checked_error_cases,
        "declared_changed_cells_checked": checked_changed_cells,
        "mutation_ledger_records": len(mutation_log),
        "mutation_ledger_exact_match": ledger == declared_ledger,
        "undeclared_formula_differences": undeclared_formula_differences,
        "nonformula_differences": nonformula_differences,
        "sheet_structure_differences": sheet_structure_differences,
        "secret_original_workbooks": sum(
            name.startswith("SECRET/originals/") for name in secret_names
        ),
    }
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if not result["passed"]:
        raise SystemExit("release audit failed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
