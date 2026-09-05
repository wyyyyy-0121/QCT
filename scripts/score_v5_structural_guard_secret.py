"""Score an immutable V5 Structural Guard RECALC prediction lock once."""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import stat
import statistics
import sys
import zipfile
from collections import Counter
from collections.abc import Mapping, Sequence
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.run_v5_structural_guard_public_predictions import (
    LOCKED_MODEL_COMMIT,
    LOCKED_MODEL_FILES,
    combined_shards_sha256,
    sha256,
)

LABEL_FIELDS = (
    "case_id",
    "cluster_id",
    "identity",
    "error_type",
    "error_cells",
    "correct_formulas",
    "group_id",
    "group_members",
    "should_abstain",
    "control_reason",
    "changed_formula_count",
    "unsupported_stress",
)
ADJUDICATION_FIELDS = (
    "case_id",
    "custodian_decision",
    "reason",
    "key_value_before",
    "key_value_after",
    "office_recalc_status",
)
ERROR_TYPES = (
    "singleton",
    "contiguous_block",
    "systematic_column",
    "ambiguous_insufficient_context",
    "ambiguous_tied_templates",
)
TOP_K = (1, 3, 5, 10, 20)


def canonical_json(payload: object) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n").encode()


def write_json(path: Path, payload: object) -> None:
    path.write_bytes(canonical_json(payload))


def canonical_cell(value: str) -> tuple[str, str]:
    sheet, address = value.rsplit("!", 1)
    return sheet.strip("'"), address.replace("$", "").upper()


def formula_key(value: str | None) -> str | None:
    """Normalize whitespace/case and Excel-optional quotes around simple sheet names."""
    if value is None:
        return None
    source = value.upper()
    compact_parts: list[str] = []
    quote: str | None = None
    index = 0
    while index < len(source):
        character = source[index]
        if quote is not None and character == quote:
            if index + 1 < len(source) and source[index + 1] == quote:
                compact_parts.extend((character, character))
                index += 2
                continue
            quote = None
            compact_parts.append(character)
        elif quote is None and character in {"'", '"'}:
            quote = character
            compact_parts.append(character)
        elif not character.isspace() or quote is not None:
            compact_parts.append(character)
        index += 1
    compact = "".join(compact_parts)

    def unquote(match: re.Match[str]) -> str:
        sheet = match.group(1)
        simple = sheet.replace(".", "_").isidentifier()
        return f"{sheet}!" if simple else match.group(0)

    return re.sub(r"'([^']+)'!", unquote, compact)


def average_precision(
    ranking: Sequence[tuple[str, str]], truth: set[tuple[str, str]]
) -> float | None:
    if not truth:
        return None
    hits = 0
    total = 0.0
    for rank, cell in enumerate(ranking, 1):
        if cell in truth:
            hits += 1
            total += hits / rank
    return total / len(truth)


def csv_rows(archive: zipfile.ZipFile, name: str, fields: Sequence[str]) -> list[dict[str, str]]:
    text = archive.read(name).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if tuple(reader.fieldnames or ()) != tuple(fields):
        raise ValueError(f"unexpected fields in {name}: {reader.fieldnames}")
    return list(reader)


def verify_model_files(expected: Mapping[str, str]) -> None:
    for relative in LOCKED_MODEL_FILES:
        current = ROOT / relative
        if sha256(current) != expected.get(relative):
            raise ValueError(f"model file differs from locked metadata: {relative}")


def verify_prediction_lock(
    locked: Path, public_root: Path, public_archive: Path
) -> tuple[list[dict[str, object]], dict[str, object], dict[str, object]]:
    lock_path = locked / "prediction_lock.json"
    metadata_path = locked / "prediction_metadata.json"
    summary_path = locked / "unlabeled_summary.json"
    lock = json.loads(lock_path.read_text(encoding="utf-8"))
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    json.loads(summary_path.read_text(encoding="utf-8"))
    if not lock.get("locked") or lock.get("cases") != 360:
        raise ValueError("prediction lock is missing or incomplete")
    if lock.get("evidence_scope") != "label_free_public_recalc_prediction":
        raise ValueError("prediction lock is not the RECALC lock")
    if lock.get("labels_read") != [] or metadata.get("labels_read") != []:
        raise ValueError("prediction metadata does not preserve the label-free boundary")
    if sha256(metadata_path) != lock.get("metadata_sha256"):
        raise ValueError("prediction metadata changed after locking")
    if sha256(summary_path) != lock.get("summary_sha256"):
        raise ValueError("prediction summary changed after locking")
    if sha256(public_root / "manifest.csv") != metadata.get("manifest_sha256"):
        raise ValueError("PUBLIC manifest differs from locked metadata")
    if sha256(public_archive) != metadata.get("public_archive_sha256"):
        raise ValueError("PUBLIC archive differs from locked metadata")
    if metadata.get("model_origin_commit") != LOCKED_MODEL_COMMIT:
        raise ValueError("unexpected model origin commit")
    verify_model_files(metadata.get("model_source_sha256", {}))
    shard_paths = sorted((locked / "shards").glob("*.json"))
    if len(shard_paths) != 360:
        raise ValueError("prediction shard set is incomplete")
    if combined_shards_sha256(shard_paths) != lock.get("combined_shards_sha256"):
        raise ValueError("prediction shards changed after locking")
    shards = [json.loads(path.read_text(encoding="utf-8")) for path in shard_paths]
    if {str(row["case_id"]) for row in shards} != {path.stem for path in shard_paths}:
        raise ValueError("prediction shard identities are inconsistent")
    return shards, lock, metadata


def validate_secret_archive(
    secret_path: Path, public_root: Path
) -> tuple[dict[str, dict[str, str]], dict[str, object]]:
    with (public_root / "manifest.csv").open("r", encoding="utf-8-sig", newline="") as stream:
        manifest_rows = list(csv.DictReader(stream))
    manifest = {row["case_id"]: row for row in manifest_rows}
    with zipfile.ZipFile(secret_path) as archive:
        unsafe = [
            name
            for name in archive.namelist()
            if PurePosixPath(name).is_absolute()
            or ".." in PurePosixPath(name).parts
            or "\\" in name
        ]
        symlinks = [
            item.filename
            for item in archive.infolist()
            if stat.S_IFMT(item.external_attr >> 16) == stat.S_IFLNK
        ]
        if unsafe or symlinks:
            raise ValueError("SECRET archive contains unsafe paths or symbolic links")
        labels = csv_rows(archive, "SECRET/labels.csv", LABEL_FIELDS)
        adjudications = csv_rows(
            archive, "SECRET/adjudication.csv", ADJUDICATION_FIELDS
        )
        mutations = [
            json.loads(line)
            for line in archive.read("SECRET/mutation_log.jsonl")
            .decode("utf-8-sig")
            .splitlines()
            if line
        ]
        cluster_rows = list(
            csv.DictReader(
                io.StringIO(
                    archive.read("SECRET/provenance_and_license/clusters.csv").decode(
                        "utf-8-sig"
                    )
                )
            )
        )
        label_by_id = {row["case_id"]: row for row in labels}
        adjudication_by_id = {row["case_id"]: row for row in adjudications}
        mutation_by_id = {str(row["case_id"]): row for row in mutations}
        originals = {
            Path(name).stem: name
            for name in archive.namelist()
            if name.startswith("SECRET/originals/") and name.endswith(".xlsx")
        }
        errors = {case_id for case_id, row in label_by_id.items() if row["identity"] == "error"}
        controls = {case_id for case_id, row in label_by_id.items() if row["identity"] == "control"}
        if not (
            len(labels) == len(label_by_id) == 360
            and len(adjudications) == len(adjudication_by_id) == 360
            and len(mutations) == len(mutation_by_id) == 240
            and len(errors) == 240
            and len(controls) == 120
            and len(cluster_rows) == len({row["cluster_id"] for row in cluster_rows}) == 30
        ):
            raise ValueError("SECRET cohort counts or identifiers are invalid")
        if set(label_by_id) != set(adjudication_by_id) or set(label_by_id) != set(manifest):
            raise ValueError("SECRET and PUBLIC case identifiers differ")
        if set(mutation_by_id) != errors or set(originals) != errors:
            raise ValueError("SECRET mutations/originals do not match error cases")
        if {row["cluster_id"] for row in cluster_rows} != {
            row["cluster_id"] for row in manifest_rows
        }:
            raise ValueError("SECRET and PUBLIC clusters differ")
        if any(
            label_by_id[case_id]["cluster_id"] != manifest[case_id]["cluster_id"]
            for case_id in manifest
        ):
            raise ValueError("SECRET case-to-cluster mapping differs from PUBLIC")

        changed_formula_total = 0
        for case_id in sorted(errors):
            label = label_by_id[case_id]
            mutation = mutation_by_id[case_id]
            label_cells = set(json.loads(label["error_cells"]))
            expected = {
                item["cell"]: item["formula"]
                for item in json.loads(label["correct_formulas"])
            }
            logged = {
                f"{mutation['worksheet']}!{item['cell']}": item
                for item in mutation["changed_cells"]
            }
            if not label_cells == set(expected) == set(logged):
                raise ValueError(f"SECRET error-cell ledgers differ: {case_id}")
            if int(label["changed_formula_count"]) != len(label_cells):
                raise ValueError(f"SECRET changed-formula count differs: {case_id}")
            if mutation["error_type"] != label["error_type"]:
                raise ValueError(f"SECRET error type differs: {case_id}")
            changed_formula_total += len(label_cells)
            original = load_workbook(
                io.BytesIO(archive.read(originals[case_id])), data_only=False
            )
            public = load_workbook(
                public_root / manifest[case_id]["workbook_path"], data_only=False
            )
            if original.sheetnames != public.sheetnames:
                raise ValueError(f"original/PUBLIC sheet structure differs: {case_id}")
            observed: set[str] = set()
            for sheet_name in original.sheetnames:
                left, right = original[sheet_name], public[sheet_name]
                if (
                    left.max_row,
                    left.max_column,
                    sorted(map(str, left.merged_cells.ranges)),
                ) != (
                    right.max_row,
                    right.max_column,
                    sorted(map(str, right.merged_cells.ranges)),
                ):
                    raise ValueError(f"original/PUBLIC cell structure differs: {case_id}")
                for key in set(left._cells) | set(right._cells):
                    before = left._cells.get(key).value if key in left._cells else None
                    after = right._cells.get(key).value if key in right._cells else None
                    before_formula = isinstance(before, str) and before.startswith("=")
                    after_formula = isinstance(after, str) and after.startswith("=")
                    qualified = f"{sheet_name}!{left.cell(*key).coordinate}"
                    if before_formula or after_formula:
                        if not (
                            before_formula
                            and after_formula
                            and formula_key(before) == formula_key(after)
                        ):
                            observed.add(qualified)
                    elif before != after:
                        raise ValueError(f"non-formula value differs from original: {case_id}")
            if observed != label_cells:
                raise ValueError(f"observed mutations differ from labels: {case_id}")
            for cell, correct in expected.items():
                sheet, address = canonical_cell(cell)
                if formula_key(original[sheet][address].value) != formula_key(correct):
                    raise ValueError(f"original formula differs from label: {case_id} {cell}")
                if formula_key(public[sheet][address].value) != formula_key(
                    logged[cell]["mutated_formula"]
                ):
                    raise ValueError(f"PUBLIC formula differs from mutation log: {case_id} {cell}")
            original.close()
            public.close()
    audit = {
        "cases": 360,
        "errors": 240,
        "controls": 120,
        "clusters": 30,
        "mutation_rows": 240,
        "original_workbooks": 240,
        "changed_formula_cells": changed_formula_total,
        "error_type_counts": dict(sorted(Counter(row["error_type"] for row in labels).items())),
        "custodian_decision_counts": dict(
            sorted(Counter(row["custodian_decision"] for row in adjudications).items())
        ),
        "secret_precommit_available_in_public": False,
    }
    return label_by_id, audit


def score_case(shard: Mapping[str, object], label: Mapping[str, str]) -> dict[str, object]:
    truth = {canonical_cell(cell) for cell in json.loads(label["error_cells"])}
    expected = {
        canonical_cell(item["cell"]): item["formula"]
        for item in json.loads(label["correct_formulas"])
    }
    ranking_rows = shard["ranking"]
    if not isinstance(ranking_rows, list):
        raise TypeError(f"ranking is missing: {shard['case_id']}")
    ranking = [(str(row["sheet"]), str(row["cell"]).upper()) for row in ranking_rows]
    if len(ranking) != len(set(ranking)) or len(ranking) != int(shard["formula_count"]):
        raise ValueError(f"ranking is incomplete: {shard['case_id']}")
    candidates = {
        (str(row["sheet"]), str(row["cell"]).upper()): str(row["candidate_formula"])
        for row in ranking_rows
        if row.get("candidate_formula") is not None
    }
    candidate_truth = set(candidates) & truth
    exact = sum(formula_key(candidates[cell]) == formula_key(expected[cell]) for cell in candidate_truth)
    group_candidates = {
        cell: formula
        for cell, formula in candidates.items()
        if next(
            row
            for row in ranking_rows
            if (str(row["sheet"]), str(row["cell"]).upper()) == cell
        )["evidence"].get("group_propagated")
    }
    groups = {
        str(row["evidence"].get("group_id")): (
            str(row["evidence"].get("group_state")),
            str(row["evidence"].get("group_reason")),
        )
        for row in ranking_rows
        if row["evidence"].get("group_id")
    }
    top_hits = {str(k): len(set(ranking[:k]) & truth) for k in TOP_K}
    return {
        "case_id": shard["case_id"],
        "cluster_id": shard["cluster_id"],
        "identity": label["identity"],
        "error_type": label["error_type"],
        "should_abstain": label["should_abstain"] == "true",
        "unsupported_stress": label["unsupported_stress"] == "true",
        "formula_cells": len(ranking),
        "error_cells": len(truth),
        "average_precision": average_precision(ranking, truth),
        "top_hits": top_hits,
        "candidate_cells": len(candidates),
        "candidate_truth_hits": len(candidate_truth),
        "candidate_exact_repairs": exact,
        "group_candidate_cells": len(group_candidates),
        "group_exact_repairs": sum(
            cell in expected and formula_key(formula) == formula_key(expected[cell])
            for cell, formula in group_candidates.items()
        ),
        "accepted_groups": sum(state == "accepted" for state, _ in groups.values()),
        "abstained_groups": sum(state == "abstained" for state, _ in groups.values()),
    }


def summarize(rows: Sequence[Mapping[str, object]]) -> dict[str, object]:
    truth = sum(int(row["error_cells"]) for row in rows)
    formulas = sum(int(row["formula_cells"]) for row in rows)
    candidates = sum(int(row["candidate_cells"]) for row in rows)
    candidate_truth = sum(int(row["candidate_truth_hits"]) for row in rows)
    exact = sum(int(row["candidate_exact_repairs"]) for row in rows)
    aps = [float(row["average_precision"]) for row in rows if row["average_precision"] is not None]
    top_hits = {
        str(k): sum(int(row["top_hits"][str(k)]) for row in rows)  # type: ignore[index]
        for k in TOP_K
    }
    return {
        "cases": len(rows),
        "formula_cells": formulas,
        "error_cells": truth,
        "macro_average_precision": statistics.fmean(aps) if aps else None,
        "top_hits": top_hits,
        "top_recall": {key: value / truth if truth else None for key, value in top_hits.items()},
        "cases_with_top5_hit": sum(int(row["top_hits"]["5"]) > 0 for row in rows),  # type: ignore[index]
        "candidate_cells": candidates,
        "cases_with_candidates": sum(int(row["candidate_cells"]) > 0 for row in rows),
        "candidate_truth_hits": candidate_truth,
        "candidate_location_precision": candidate_truth / candidates if candidates else None,
        "candidate_error_coverage": candidate_truth / truth if truth else None,
        "candidate_exact_repairs": exact,
        "candidate_exact_precision": exact / candidates if candidates else None,
        "candidate_exact_coverage": exact / truth if truth else None,
        "group_candidate_cells": sum(int(row["group_candidate_cells"]) for row in rows),
        "group_exact_repairs": sum(int(row["group_exact_repairs"]) for row in rows),
        "accepted_groups": sum(int(row["accepted_groups"]) for row in rows),
        "abstained_groups": sum(int(row["abstained_groups"]) for row in rows),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--locked", type=Path, required=True)
    parser.add_argument("--public", type=Path, required=True)
    parser.add_argument("--public-archive", type=Path, required=True)
    parser.add_argument("--secret-archive", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    shards, lock, metadata = verify_prediction_lock(
        args.locked.resolve(), args.public.resolve(), args.public_archive.resolve()
    )
    labels, secret_audit = validate_secret_archive(
        args.secret_archive.resolve(), args.public.resolve()
    )
    if {str(row["case_id"]) for row in shards} != set(labels):
        raise SystemExit("scoring refused: prediction and SECRET case IDs differ")
    rows = [score_case(shard, labels[str(shard["case_id"])]) for shard in shards]
    by_error_type = {
        error_type: summarize([row for row in rows if row["error_type"] == error_type])
        for error_type in (*ERROR_TYPES, "none")
    }
    error_rows = [row for row in rows if row["identity"] == "error"]
    control_rows = [row for row in rows if row["identity"] == "control"]
    repair_rows = [
        row for row in error_rows if not row["should_abstain"]
    ]
    abstain_rows = [row for row in rows if row["should_abstain"]]
    receipt = {
        "protocol": "v5_structural_guard_r2_recalc_secret_scoring_v1",
        "outcome": "descriptive_scoring_complete_no_preregistered_promotion_rule",
        "promotion_decision": None,
        "promotion_decision_reason": (
            "No cohort-specific promotion thresholds or SECRET archive hash were committed "
            "with the PUBLIC package before SECRET release."
        ),
        "post_secret_tuning_forbidden": True,
        "prediction_lock_sha256": sha256(args.locked / "prediction_lock.json"),
        "combined_shards_sha256": lock["combined_shards_sha256"],
        "prediction_runner_commit": metadata["prediction_runner_commit"],
        "public_archive_sha256": sha256(args.public_archive),
        "secret_archive_sha256": sha256(args.secret_archive),
        "secret_audit": secret_audit,
        "overall_error_cases": summarize(error_rows),
        "controls": summarize(control_rows),
        "repair_expected_cases": summarize(repair_rows),
        "should_abstain_cases": summarize(abstain_rows),
        "by_error_type": by_error_type,
    }
    args.output.mkdir(parents=True, exist_ok=False)
    write_json(args.output / "scoring_receipt.json", receipt)
    with (args.output / "case_scores.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        flat_rows = []
        for row in rows:
            flat = {key: value for key, value in row.items() if key != "top_hits"}
            flat.update({f"top_{key}_hits": value for key, value in row["top_hits"].items()})  # type: ignore[union-attr]
            flat_rows.append(flat)
        writer = csv.DictWriter(stream, fieldnames=list(flat_rows[0]))
        writer.writeheader()
        writer.writerows(flat_rows)
    print(json.dumps(receipt, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
