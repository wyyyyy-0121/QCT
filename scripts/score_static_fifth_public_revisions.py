#!/usr/bin/env python3
"""Score locked static-fifth predictions on public workbook revisions."""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import shutil
import statistics
import subprocess
import sys
import tempfile
from collections import defaultdict
from collections.abc import Mapping, Sequence
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from formulaguard.formula import normalized_formula
from scripts.run_static_fifth_revision_predictions import (
    EXPECTED_ARCHIVE_SHA256,
    REVIEW_BUDGET,
    canonical_json,
    sha256,
    sha256_bytes,
    stable_hash,
)
from scripts.run_static_fifth_revision_predictions import (
    PROTOCOL as PREDICTION_PROTOCOL,
)

PROTOCOL = "formulaguard_static_fifth_public_revision_score_v1"
EXPECTED_REVISIONS = 4
EXPECTED_FORMULA_CHANGES = 8
MANIFEST_MEMBER = "public_revisions/manifest.json"
CASES_MEMBER = "public_revisions/cases.csv"
CASE_FIELDS = (
    "revision_case_id",
    "workbook_id",
    "repository",
    "commit_sha",
    "parent_sha",
    "committed_at",
    "commit_message",
    "workbook_path",
    "sheet_name",
    "cell_address",
    "before_formula",
    "after_formula",
    "license_spdx",
    "commit_url",
    "before_url",
    "after_url",
    "before_sha256",
    "after_sha256",
)
SOURCE_PATHS = (
    "formulaguard/formula.py",
    "scripts/run_static_fifth_revision_predictions.py",
    "scripts/score_static_fifth_public_revisions.py",
)

DEFAULT_ARCHIVE = Path(
    "/home/ayaka/code/FormulaGuard_public_revisions_delivery_20260831.zip"
)
DEFAULT_RUN_A = ROOT / "results/static_fifth_revision_predictions_run_a"
DEFAULT_RUN_B = ROOT / "results/static_fifth_revision_predictions_run_b"
DEFAULT_OUTPUT = ROOT / "results/static_fifth_public_revision_score"


def git_commit(root: Path = ROOT) -> str:
    return subprocess.check_output(
        ("git", "rev-parse", "HEAD"), cwd=root, text=True
    ).strip()


def _git_source_status(root: Path) -> tuple[str, ...]:
    completed = subprocess.run(
        ("git", "status", "--porcelain", "--", *SOURCE_PATHS),
        cwd=root,
        check=True,
        capture_output=True,
        text=True,
    )
    return tuple(line for line in completed.stdout.splitlines() if line)


def capture_source_state(
    source_root: Path = ROOT,
    *,
    allow_dirty: bool = False,
) -> dict[str, object]:
    source_root = source_root.resolve()
    state = {
        "git_commit": git_commit(source_root),
        "source_sha256": {
            relative: sha256(source_root / relative) for relative in SOURCE_PATHS
        },
        "source_status": list(_git_source_status(source_root)),
    }
    dirty = bool(state["source_status"])
    if dirty and not allow_dirty:
        raise ValueError("formal scoring requires clean tracked scorer sources")
    state["source_tree_dirty"] = dirty
    state["formal_evidence"] = not dirty
    return state


def verify_source_state(
    expected: Mapping[str, object], source_root: Path = ROOT
) -> None:
    observed = capture_source_state(source_root, allow_dirty=True)
    if any(
        observed[key] != expected[key]
        for key in ("git_commit", "source_sha256", "source_status")
    ):
        raise ValueError("scorer source changed during scoring")


def _load_json(path: Path) -> dict[str, object]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise TypeError(f"JSON object required: {path}")
    return value


def validate_prediction_run(run_dir: Path) -> dict[str, object]:
    if not run_dir.resolve().is_dir():
        raise ValueError(f"prediction run is not a directory: {run_dir}")
    receipt_path = run_dir / "completion_receipt.json"
    metadata_path = run_dir / "prediction_metadata.json"
    predictions_path = run_dir / "predictions.jsonl"
    receipt = _load_json(receipt_path)
    metadata = _load_json(metadata_path)
    if receipt.get("protocol") != PREDICTION_PROTOCOL:
        raise ValueError("prediction receipt protocol mismatch")
    if receipt.get("complete") is not True or receipt.get("formal_evidence") is not True:
        raise ValueError("prediction run is not complete formal evidence")
    for payload in (receipt, metadata):
        if payload.get("label_members_read") != [] or payload.get("label_inputs") != []:
            raise ValueError("prediction run declares label input")
        if payload.get("protected_data_inputs") != []:
            raise ValueError("prediction run declares protected input")
        if payload.get("archive_sha256") != EXPECTED_ARCHIVE_SHA256:
            raise ValueError("prediction archive hash mismatch")
    expected_before_members = [
        f"public_revisions/workbooks/PWR{index:03d}/before.xlsx"
        for index in range(1, EXPECTED_REVISIONS + 1)
    ]
    if metadata.get("archive_payload_members_read") != expected_before_members:
        raise ValueError("prediction read an unexpected archive payload member")
    if receipt.get("prediction_metadata_sha256") != sha256(metadata_path):
        raise ValueError("prediction metadata hash mismatch")
    if receipt.get("predictions_sha256") != sha256(predictions_path):
        raise ValueError("merged prediction hash mismatch")
    shard_hashes = receipt.get("prediction_shard_sha256")
    if not isinstance(shard_hashes, Mapping) or len(shard_hashes) != EXPECTED_REVISIONS:
        raise TypeError("prediction shard inventory is malformed")
    if receipt.get("shard_inventory_sha256") != stable_hash(shard_hashes):
        raise ValueError("prediction shard inventory hash mismatch")
    expected_paths = {
        (run_dir / str(relative)).resolve(): str(digest)
        for relative, digest in shard_hashes.items()
    }
    observed_paths = set((run_dir / "shards").glob("*.json"))
    if set(expected_paths) != observed_paths:
        raise ValueError("prediction shard files differ from the receipt")
    records: list[dict[str, object]] = []
    seen: set[str] = set()
    for path in sorted(observed_paths, key=lambda item: item.name):
        if sha256(path) != expected_paths[path]:
            raise ValueError(f"prediction shard hash mismatch: {path.name}")
        record = _load_json(path)
        if record.get("protocol") != PREDICTION_PROTOCOL:
            raise ValueError(f"prediction shard protocol mismatch: {path.name}")
        if record.get("label_members_read") != [] or record.get("label_inputs") != []:
            raise ValueError(f"prediction shard declares label input: {path.name}")
        if record.get("protected_data_inputs") != []:
            raise ValueError(f"prediction shard declares protected input: {path.name}")
        revision_id = str(record.get("revision_id", ""))
        if path.name != f"{revision_id}.json" or revision_id in seen:
            raise ValueError(f"prediction shard identity mismatch: {path.name}")
        seen.add(revision_id)
        rankings = record.get("rankings")
        if not isinstance(rankings, Mapping) or set(rankings) != {
            "v4_r1",
            "v4_static_fifth",
        }:
            raise ValueError(f"prediction ranking methods are malformed: {path.name}")
        formula_count = int(record.get("formula_count", -1))
        inventories: list[set[str]] = []
        for method, raw_ranking in rankings.items():
            if not isinstance(raw_ranking, list) or len(raw_ranking) != formula_count:
                raise ValueError(f"{method} ranking is incomplete: {path.name}")
            cells = [str(item["cell"]) for item in raw_ranking]
            ranks = [int(item["rank"]) for item in raw_ranking]
            if ranks != list(range(1, formula_count + 1)):
                raise ValueError(f"{method} ranks are not contiguous: {path.name}")
            if len(cells) != len(set(cells)):
                raise ValueError(f"{method} ranking has duplicate cells: {path.name}")
            inventories.append(set(cells))
        if inventories[0] != inventories[1]:
            raise ValueError("V4 and candidate formula inventories differ")
        v4_top4 = [str(item["cell"]) for item in rankings["v4_r1"][:4]]
        candidate_top4 = [
            str(item["cell"]) for item in rankings["v4_static_fifth"][:4]
        ]
        if v4_top4 != candidate_top4 or record.get("review_budget") != REVIEW_BUDGET:
            raise ValueError("prediction changed the frozen comparison budget or V4 Top-4")
        records.append(record)
    records.sort(key=lambda item: str(item["revision_id"]))
    if receipt.get("record_set_sha256") != stable_hash(records):
        raise ValueError("prediction record-set hash mismatch")
    expected_jsonl = "".join(canonical_json(record) + "\n" for record in records)
    if predictions_path.read_text(encoding="ascii") != expected_jsonl:
        raise ValueError("merged predictions differ from the locked shards")
    return {
        "receipt": receipt,
        "metadata": metadata,
        "records": records,
        "records_by_revision": {
            str(record["revision_id"]): record for record in records
        },
        "files": sorted(
            path.relative_to(run_dir.resolve()).as_posix()
            for path in run_dir.resolve().rglob("*")
            if path.is_file()
        ),
    }


def validate_prediction_pair(
    run_a: Path, run_b: Path
) -> tuple[dict[str, object], dict[str, object]]:
    first = validate_prediction_run(run_a)
    second = validate_prediction_run(run_b)
    if first["files"] != second["files"]:
        raise ValueError("prediction runs have different file sets")
    for relative in first["files"]:
        if (run_a / str(relative)).read_bytes() != (run_b / str(relative)).read_bytes():
            raise ValueError(f"prediction runs differ: {relative}")
    return first, second


def _formula_map(payload: bytes, directory: Path, name: str) -> dict[str, str]:
    path = directory / name
    path.write_bytes(payload)
    workbook = load_workbook(path, read_only=True, data_only=False)
    formulas: dict[str, str] = {}
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas[f"{sheet.title}!{cell.coordinate}"] = value
    finally:
        workbook.close()
    return formulas


def load_revision_labels(
    archive: Path,
    *,
    prediction_records: Mapping[str, object],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    archive = archive.resolve()
    if "FormulaGuard_240_120" in archive.parts:
        raise ValueError("protected 240+120 path is forbidden")
    if sha256(archive) != EXPECTED_ARCHIVE_SHA256:
        raise ValueError("public revision archive hash changed before scoring")
    try:
        with ZipFile(archive) as source:
            manifest_bytes = source.read(MANIFEST_MEMBER)
            cases_bytes = source.read(CASES_MEMBER)
            manifest = json.loads(manifest_bytes.decode("utf-8"))
            if not isinstance(manifest, Mapping):
                raise TypeError("public revision manifest must be an object")
            if manifest.get("protocol") != "formulaguard_public_xlsx_revision_evidence_v1":
                raise ValueError("public revision manifest protocol mismatch")
            counts = manifest.get("counts")
            if not isinstance(counts, Mapping):
                raise TypeError("public revision manifest counts are missing")
            if counts.get("workbook_pairs") != EXPECTED_REVISIONS:
                raise ValueError("public revision workbook-pair count mismatch")
            if counts.get("formula_revision_cells") != EXPECTED_FORMULA_CHANGES:
                raise ValueError("public revision formula-change count mismatch")
            workbooks = manifest.get("workbooks")
            if not isinstance(workbooks, list) or len(workbooks) != EXPECTED_REVISIONS:
                raise ValueError("public revision workbook inventory is malformed")

            with io.StringIO(cases_bytes.decode("utf-8-sig"), newline="") as handle:
                reader = csv.DictReader(handle)
                if tuple(reader.fieldnames or ()) != CASE_FIELDS:
                    raise ValueError("public revision case schema mismatch")
                cases = list(reader)
            if len(cases) != EXPECTED_FORMULA_CHANGES:
                raise ValueError("public revision case count mismatch")
            if len({row["revision_case_id"] for row in cases}) != len(cases):
                raise ValueError("duplicate public revision case IDs")

            case_groups: dict[str, list[dict[str, str]]] = defaultdict(list)
            for row in cases:
                case_groups[row["workbook_id"]].append(row)
            events: list[dict[str, object]] = []
            after_inventory: dict[str, str] = {}
            label_members = [MANIFEST_MEMBER, CASES_MEMBER]
            with tempfile.TemporaryDirectory(prefix="public-revision-score-") as temp:
                directory = Path(temp)
                for workbook in sorted(workbooks, key=lambda item: str(item["workbook_id"])):
                    revision_id = str(workbook["workbook_id"])
                    prediction = prediction_records.get(revision_id)
                    if not isinstance(prediction, Mapping):
                        raise TypeError(f"missing locked prediction: {revision_id}")
                    before_member = "public_revisions/" + str(workbook["before_file"])
                    after_member = "public_revisions/" + str(workbook["after_file"])
                    before_bytes = source.read(before_member)
                    after_bytes = source.read(after_member)
                    label_members.extend((before_member, after_member))
                    before_hash = sha256_bytes(before_bytes)
                    after_hash = sha256_bytes(after_bytes)
                    if before_hash != workbook["before_sha256"]:
                        raise ValueError(f"before workbook hash mismatch: {revision_id}")
                    if after_hash != workbook["after_sha256"]:
                        raise ValueError(f"after workbook hash mismatch: {revision_id}")
                    if prediction.get("workbook_sha256") != before_hash:
                        raise ValueError(f"prediction/workbook hash mismatch: {revision_id}")
                    before_formulas = _formula_map(
                        before_bytes, directory, f"{revision_id}-before.xlsx"
                    )
                    after_formulas = _formula_map(
                        after_bytes, directory, f"{revision_id}-after.xlsx"
                    )
                    diff = {
                        cell: (before_formulas.get(cell), after_formulas.get(cell))
                        for cell in set(before_formulas) | set(after_formulas)
                        if before_formulas.get(cell) != after_formulas.get(cell)
                    }
                    declared = {
                        f"{row['sheet_name']}!{row['cell_address']}": (
                            row["before_formula"],
                            row["after_formula"],
                        )
                        for row in case_groups[revision_id]
                    }
                    if diff != declared:
                        raise ValueError(f"independent formula diff mismatch: {revision_id}")
                    if len(declared) != int(workbook["formula_change_count"]):
                        raise ValueError(f"formula-change count mismatch: {revision_id}")
                    for row in case_groups[revision_id]:
                        if row["before_sha256"] != before_hash or row["after_sha256"] != after_hash:
                            raise ValueError(f"case workbook hash mismatch: {revision_id}")
                        if row["license_spdx"] != "MIT":
                            raise ValueError(f"case license mismatch: {revision_id}")
                    after_inventory[after_member] = after_hash
                    events.append(
                        {
                            "revision_id": revision_id,
                            "commit_sha": workbook["commit_sha"],
                            "before_workbook_sha256": before_hash,
                            "after_workbook_sha256": after_hash,
                            "source_cells": sorted(declared),
                            "formula_change_count": len(declared),
                            "after_formulas": {
                                cell: declared[cell][1] for cell in sorted(declared)
                            },
                        }
                    )
    except BadZipFile as exc:
        raise ValueError("public revision archive is not a valid ZIP") from exc
    if {event["revision_id"] for event in events} != set(prediction_records):
        raise ValueError("revealed revision IDs differ from locked predictions")
    return events, {
        "archive": archive.as_posix(),
        "archive_sha256": EXPECTED_ARCHIVE_SHA256,
        "manifest_member": MANIFEST_MEMBER,
        "manifest_sha256": sha256_bytes(manifest_bytes),
        "cases_member": CASES_MEMBER,
        "cases_sha256": sha256_bytes(cases_bytes),
        "after_workbook_inventory_sha256": stable_hash(after_inventory),
        "label_payload_members_read": label_members,
        "workbook_pairs": len(events),
        "formula_revision_cells": sum(
            int(event["formula_change_count"]) for event in events
        ),
        "protected_data_inputs": [],
    }


def ranking_metric(
    ranking: Sequence[Mapping[str, object]], source_cells: Sequence[str]
) -> dict[str, object]:
    positions = {str(item["cell"]): int(item["rank"]) for item in ranking}
    ranks = [positions[cell] for cell in source_cells if cell in positions]
    rank = min(ranks) if ranks else None
    return {
        "rank": rank,
        "top1": int(rank is not None and rank <= 1),
        "top5": int(rank is not None and rank <= REVIEW_BUDGET),
        "mrr": 1.0 / rank if rank is not None else 0.0,
        "formula_cells_top5": sum(
            positions.get(cell, REVIEW_BUDGET + 1) <= REVIEW_BUDGET
            for cell in source_cells
        ),
    }


def evaluate_gates(
    *,
    integrity_passed: bool,
    revision_events: int,
    formula_changes: int,
    inventory_parity: bool,
    net_top5_rescues: int,
    candidate_top5: float,
    v4_top5: float,
    top5_losses: int,
    per_revision_regressions: int,
    candidate_mrr: float,
    v4_mrr: float,
    protected_data_inputs: Sequence[str],
) -> dict[str, object]:
    gates = {
        "g1_prediction_and_scoring_integrity": integrity_passed,
        "g2_complete_revision_evidence": revision_events == EXPECTED_REVISIONS
        and formula_changes == EXPECTED_FORMULA_CHANGES,
        "g3_same_inventory_and_review_budget": inventory_parity,
        "g4_strict_top5_improvement": net_top5_rescues >= 1
        and candidate_top5 > v4_top5,
        "g5_zero_losses_and_regressions": top5_losses == 0
        and per_revision_regressions == 0,
        "g6_mrr_non_degradation": candidate_mrr >= v4_mrr,
        "g7_no_protected_input": not protected_data_inputs,
    }
    return {
        **gates,
        "all_gates_passed": all(gates.values()),
        "failed_gates": sorted(key for key, value in gates.items() if not value),
    }


def build_scores(
    events: Sequence[Mapping[str, object]],
    predictions: Mapping[str, object],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    rows: list[dict[str, object]] = []
    inventory_parity = True
    for event in events:
        revision_id = str(event["revision_id"])
        prediction = predictions.get(revision_id)
        if not isinstance(prediction, Mapping):
            raise TypeError(f"locked prediction is malformed: {revision_id}")
        rankings = prediction.get("rankings")
        if not isinstance(rankings, Mapping):
            raise TypeError(f"locked rankings are malformed: {revision_id}")
        v4_ranking = rankings["v4_r1"]
        candidate_ranking = rankings["v4_static_fifth"]
        if not isinstance(v4_ranking, list) or not isinstance(candidate_ranking, list):
            raise TypeError(f"locked ranking rows are malformed: {revision_id}")
        v4_cells = {str(item["cell"]) for item in v4_ranking}
        candidate_cells = {str(item["cell"]) for item in candidate_ranking}
        inventory_parity &= v4_cells == candidate_cells
        sources = [str(cell) for cell in event["source_cells"]]
        v4_metric = ranking_metric(v4_ranking, sources)
        candidate_metric = ranking_metric(candidate_ranking, sources)
        selected = candidate_ranking[REVIEW_BUDGET - 1]
        selected_cell = str(selected["cell"])
        after_formulas = event["after_formulas"]
        if not isinstance(after_formulas, Mapping):
            raise TypeError("revealed after formulas are malformed")
        selected_formula = selected.get("candidate_formula")
        repair_exact = (
            selected_cell in after_formulas
            and isinstance(selected_formula, str)
            and normalized_formula(selected_formula)
            == normalized_formula(str(after_formulas[selected_cell]))
        )
        rows.append(
            {
                "revision_id": revision_id,
                "commit_sha": event["commit_sha"],
                "before_workbook_sha256": event["before_workbook_sha256"],
                "after_workbook_sha256": event["after_workbook_sha256"],
                "source_cells": sources,
                "formula_change_count": event["formula_change_count"],
                "static_fifth_cell": selected_cell,
                "static_fifth_candidate_formula": selected_formula,
                "static_fifth_repair_exact_if_source": repair_exact,
                "metrics": {
                    "v4_r1": v4_metric,
                    "v4_static_fifth": candidate_metric,
                },
            }
        )
    rows.sort(key=lambda item: str(item["revision_id"]))
    rescues = [
        row
        for row in rows
        if row["metrics"]["v4_r1"]["top5"] == 0  # type: ignore[index]
        and row["metrics"]["v4_static_fifth"]["top5"] == 1  # type: ignore[index]
    ]
    losses = [
        row
        for row in rows
        if row["metrics"]["v4_r1"]["top5"] == 1  # type: ignore[index]
        and row["metrics"]["v4_static_fifth"]["top5"] == 0  # type: ignore[index]
    ]
    regressions = [
        row
        for row in rows
        if int(row["metrics"]["v4_static_fifth"]["rank"])  # type: ignore[index]
        > int(row["metrics"]["v4_r1"]["rank"])  # type: ignore[index]
    ]
    methods: dict[str, dict[str, object]] = {}
    for method in ("v4_r1", "v4_static_fifth"):
        methods[method] = {
            "events": len(rows),
            "top1": statistics.fmean(
                float(row["metrics"][method]["top1"]) for row in rows  # type: ignore[index]
            ),
            "top5": statistics.fmean(
                float(row["metrics"][method]["top5"]) for row in rows  # type: ignore[index]
            ),
            "mrr": statistics.fmean(
                float(row["metrics"][method]["mrr"]) for row in rows  # type: ignore[index]
            ),
            "formula_cells_top5": sum(
                int(row["metrics"][method]["formula_cells_top5"])  # type: ignore[index]
                for row in rows
            ),
            "formula_cells": sum(int(row["formula_change_count"]) for row in rows),
        }
    paired = {
        "top5_rescues": len(rescues),
        "top5_losses": len(losses),
        "net_top5_rescues": len(rescues) - len(losses),
        "rescue_revision_ids": [str(row["revision_id"]) for row in rescues],
        "loss_revision_ids": [str(row["revision_id"]) for row in losses],
        "per_revision_regressions": len(regressions),
        "regression_revision_ids": [
            str(row["revision_id"]) for row in regressions
        ],
    }
    gates = evaluate_gates(
        integrity_passed=True,
        revision_events=len(rows),
        formula_changes=sum(int(row["formula_change_count"]) for row in rows),
        inventory_parity=inventory_parity,
        net_top5_rescues=paired["net_top5_rescues"],
        candidate_top5=float(methods["v4_static_fifth"]["top5"]),
        v4_top5=float(methods["v4_r1"]["top5"]),
        top5_losses=len(losses),
        per_revision_regressions=len(regressions),
        candidate_mrr=float(methods["v4_static_fifth"]["mrr"]),
        v4_mrr=float(methods["v4_r1"]["mrr"]),
        protected_data_inputs=(),
    )
    return rows, {
        "methods": methods,
        "paired": paired,
        "inventory_parity": inventory_parity,
        "gates": gates,
    }


def write_score(
    output: Path,
    rows: Sequence[Mapping[str, object]],
    summary: Mapping[str, object],
) -> Path:
    output = output.resolve()
    partial = output.with_name(output.name + ".partial")
    if output.exists() or partial.exists():
        raise ValueError("score output or partial output already exists")
    partial.mkdir(parents=True)
    try:
        events_path = partial / "revision_scores.jsonl"
        with events_path.open("w", encoding="ascii", newline="\n") as handle:
            for row in rows:
                handle.write(canonical_json(row) + "\n")
        payload = {
            **dict(summary),
            "revision_scores_sha256": sha256(events_path),
            "revision_score_set_sha256": stable_hash(rows),
        }
        summary_path = partial / "score_summary.json"
        summary_path.write_text(
            json.dumps(payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n",
            encoding="ascii",
        )
        os.replace(partial, output)
    except BaseException:
        shutil.rmtree(partial, ignore_errors=True)
        raise
    return output / "score_summary.json"


def score(
    *,
    archive: Path,
    run_a: Path,
    run_b: Path,
    output: Path,
    source_root: Path = ROOT,
    allow_dirty: bool = False,
) -> Path:
    source_state = capture_source_state(source_root, allow_dirty=allow_dirty)

    # The scorer must finish both prediction-lock audits before opening labels.
    first, second = validate_prediction_pair(run_a, run_b)
    prediction_records = first["records_by_revision"]
    if not isinstance(prediction_records, Mapping):
        raise TypeError("prediction record index is malformed")
    events, label_audit = load_revision_labels(
        archive,
        prediction_records=prediction_records,
    )
    rows, metrics = build_scores(events, prediction_records)
    verify_source_state(source_state, source_root)
    first_receipt = first["receipt"]
    second_receipt = second["receipt"]
    if not isinstance(first_receipt, Mapping) or not isinstance(second_receipt, Mapping):
        raise TypeError("prediction receipts are malformed")
    summary = {
        "protocol": PROTOCOL,
        "complete": True,
        **dict(source_state),
        "prediction_integrity_verified_before_labels": True,
        "prediction_runs_byte_identical": True,
        "prediction_runs": [
            {
                "path": run.as_posix(),
                "completion_receipt_sha256": sha256(
                    run / "completion_receipt.json"
                ),
                "git_commit": receipt["git_commit"],
                "record_set_sha256": receipt["record_set_sha256"],
                "shard_inventory_sha256": receipt["shard_inventory_sha256"],
            }
            for run, receipt in (
                (run_a, first_receipt),
                (run_b, second_receipt),
            )
        ],
        "label_audit": label_audit,
        "candidate": {
            "model_version": "v4-static-fifth-exploratory-v1",
            "review_budget": REVIEW_BUDGET,
            "unchanged_from_frozen_candidate": True,
            "formal_version": None,
        },
        "revision_events": len(rows),
        "formula_revision_cells": sum(
            int(row["formula_change_count"]) for row in rows
        ),
        **metrics,
        "decision": {
            "bounded_public_revision_candidate_authorized": metrics["gates"][
                "all_gates_passed"
            ],
            "formal_v5_r1_authorized": False,
            "protected_evaluation_authorized": False,
            "claim_scope": "single_project_public_revision_confirmation",
        },
        "protected_data_inputs": [],
    }
    return write_score(output, rows, summary)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--archive", type=Path, default=DEFAULT_ARCHIVE)
    parser.add_argument("--run-a", type=Path, default=DEFAULT_RUN_A)
    parser.add_argument("--run-b", type=Path, default=DEFAULT_RUN_B)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--allow-dirty", action="store_true")
    args = parser.parse_args(argv)
    try:
        result = score(
            archive=args.archive,
            run_a=args.run_a,
            run_b=args.run_b,
            output=args.output,
            allow_dirty=args.allow_dirty,
        )
    except (OSError, TypeError, ValueError, KeyError, BadZipFile) as exc:
        raise SystemExit(f"public revision scoring refused: {exc}") from exc
    print(f"Public revision score: {result}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
