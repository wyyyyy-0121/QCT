"""Build the frozen same-source StructuralGuard benchmark.

The benchmark is intentionally derived only from the five workbooks shipped in
StructuralGuard_handoff_20260903_complete.zip.  It contains one original
singleton-error case, one clean control, one coherent 20-row block error, and
one fully systematic column error per source workbook.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import zipfile
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

PROTOCOL = "structuralguard_standard_benchmark_v1"
FIXED_TIME = datetime(2000, 1, 1, tzinfo=UTC)


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def translated(formula: str, source_cell: str, target_cell: str) -> str:
    return Translator(formula, origin=source_cell).translate_formula(target_cell)


def canonical_core_xml(value: bytes) -> bytes:
    namespaces = {
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
        "dc": "http://purl.org/dc/elements/1.1/",
        "dcterms": "http://purl.org/dc/terms/",
        "xsi": "http://www.w3.org/2001/XMLSchema-instance",
    }
    for prefix, uri in namespaces.items():
        ElementTree.register_namespace(prefix, uri)
    root = ElementTree.fromstring(value)
    modified = root.find("{http://purl.org/dc/terms/}modified")
    if modified is not None:
        modified.text = "2000-01-01T00:00:00Z"
    return ElementTree.tostring(root, encoding="utf-8")


def normalize_workbook(wb) -> None:
    # Normalize metadata so rerunning this builder is byte-stable with the same
    # openpyxl version, while retaining formulas, values, styles, and dimensions.
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.properties.lastPrinted = None
    if getattr(wb, "calculation", None) is not None:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True


def zip_member(zf: zipfile.ZipFile, suffix: str) -> str:
    matches = [name for name in zf.namelist() if name.replace("\\", "/").endswith(suffix)]
    if len(matches) != 1:
        raise ValueError(f"expected one ZIP member ending {suffix!r}, found {matches}")
    return matches[0]


def load_answers(zf: zipfile.ZipFile) -> dict:
    member = zip_member(
        zf,
        "StructuralGuard_handoff_20260903/03_data_and_answers_OPEN_AFTER_TEST/blind_answer_keys.json",
    )
    return json.loads(zf.read(member).decode("utf-8"))


def error_map(entry: dict) -> dict[tuple[str, str], dict]:
    return {(str(item["sheet"]), str(item["cell"])): item for item in entry["errors"]}


def formula_cells(wb) -> list[tuple[str, str]]:
    cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cells.append((ws.title, cell.coordinate))
    return cells


def apply_clean_repairs(wb, entry: dict) -> None:
    for item in entry["errors"]:
        wb[str(item["sheet"])][str(item["cell"])].value = str(item["expected_formula"])


def mutation_record(
    wb,
    sheet: str,
    target_cell: str,
    source_cell: str,
    injected_formula: str,
    expected_formula: str,
    category: str,
    condition: str,
) -> dict:
    target = wb[sheet][target_cell]
    actual_expected = translated(expected_formula, source_cell, target_cell)
    actual_injected = translated(injected_formula, source_cell, target_cell)
    target.value = actual_injected
    return {
        "sheet": sheet,
        "cell": target_cell,
        "category": category,
        "injected_formula": actual_injected,
        "expected_formula": actual_expected,
        "source_template_cell": source_cell,
        "mutation_condition": condition,
    }


def save_case(wb, path: Path) -> str:
    normalize_workbook(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    # XLSX is a ZIP container; canonicalize member order and timestamps so
    # generated case hashes do not depend on the wall clock.
    with zipfile.ZipFile(buffer, "r") as source, zipfile.ZipFile(
        path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            content = source.read(name)
            if name == "docProps/core.xml":
                content = canonical_core_xml(content)
            target.writestr(info, content)
    return sha256_file(path)


def build_case(
    source_bytes: bytes,
    source_entry: dict,
    output_path: Path,
    condition: str,
    mutation_spec: dict | None,
) -> dict:
    wb = load_workbook(BytesIO(source_bytes), data_only=False)
    clean_wb = load_workbook(BytesIO(source_bytes), data_only=False)
    apply_clean_repairs(clean_wb, source_entry)
    errors: list[dict] = []
    if condition == "singleton":
        # Preserve the shipped formulas and labels for this queue.
        wb = load_workbook(BytesIO(source_bytes), data_only=False)
        errors = [
            {
                "sheet": str(item["sheet"]),
                "cell": str(item["cell"]),
                "category": str(item["category"]),
                "injected_formula": str(item["injected_formula"]),
                "expected_formula": str(item["expected_formula"]),
                "source_template_cell": str(item["cell"]),
                "mutation_condition": "shipped_singleton",
            }
            for item in source_entry["errors"]
        ]
    elif condition == "clean":
        wb = clean_wb
    else:
        wb = clean_wb
        sheet = str(mutation_spec["sheet"])
        source_cell = str(mutation_spec["source_cell"])
        source_item = mutation_spec["source_item"]
        for target_cell in mutation_spec["target_cells"]:
            errors.append(
                mutation_record(
                    wb,
                    sheet,
                    str(target_cell),
                    source_cell,
                    str(source_item["injected_formula"]),
                    str(source_item["expected_formula"]),
                    str(source_item["category"]),
                    condition,
                )
            )
    digest = save_case(wb, output_path)
    return {
        "case_id": output_path.stem,
        "condition": condition,
        "scenario": str(source_entry["scenario"]),
        "source_file": str(source_entry["file"]),
        "workbook": str(output_path.relative_to(output_path.parents[2])).replace("\\", "/"),
        "sha256": digest,
        "formula_cells": len(formula_cells(wb)),
        "errors": errors,
    }


def make_mutation_specs(entry: dict) -> tuple[dict, dict]:
    name = str(entry["file"])
    by_key = error_map(entry)
    if name.startswith("blind_01"):
        block_key, column_key, column = ("Sales Detail", "G37"), ("Sales Detail", "G37"), "G"
        full_rows = range(3, 223)
        block_rows = range(28, 48)
    elif name.startswith("blind_02"):
        block_key, column_key, column = ("Movements", "F40"), ("Movements", "F95"), "F"
        full_rows = range(3, 183)
        block_rows = range(31, 51)
    elif name.startswith("blind_03"):
        block_key, column_key, column = ("Projects", "F34"), ("Projects", "K145"), "K"
        full_rows = range(3, 163)
        block_rows = range(25, 45)
    elif name.startswith("blind_04"):
        block_key, column_key, column = ("Grades", "G25"), ("Grades", "I130"), "I"
        full_rows = range(3, 183)
        block_rows = range(16, 36)
    elif name.startswith("blind_05"):
        block_key, column_key, column = ("Customers", "F45"), ("Customers", "I101"), "I"
        full_rows = range(3, 203)
        block_rows = range(36, 56)
    else:
        raise ValueError(f"unrecognized source workbook {name}")
    block_item = by_key[block_key]
    column_item = by_key[column_key]
    return (
        {
            "sheet": block_key[0],
            "source_cell": block_key[1],
            "source_item": block_item,
            "target_cells": [f"{column}{row}" for row in block_rows],
        },
        {
            "sheet": column_key[0],
            "source_cell": column_key[1],
            "source_item": column_item,
            "target_cells": [f"{column}{row}" for row in full_rows],
        },
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_zip", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()
    source_zip = args.source_zip.resolve()
    output_dir = args.output_dir.resolve()
    if output_dir.exists():
        # Refresh only files owned by this builder; preserve README or other
        # user-authored notes placed beside the generated benchmark.
        for name in ("public", "labels", "metadata.json", "SHA256SUMS.txt"):
            target = output_dir / name
            if target.is_dir():
                shutil.rmtree(target)
            elif target.exists():
                target.unlink()
    (output_dir / "public" / "workbooks").mkdir(parents=True)
    (output_dir / "labels").mkdir(parents=True)

    archive_sha = sha256_file(source_zip)
    with zipfile.ZipFile(source_zip) as zf:
        answers = load_answers(zf)
        cases: list[dict] = []
        for index, entry in enumerate(answers["workbooks"], 1):
            source_name = str(entry["file"])
            source_bytes = zf.read(
                zip_member(
                    zf,
                    f"StructuralGuard_handoff_20260903/02_excel_blind_inputs/{source_name}",
                )
            )
            source_hash = sha256_bytes(source_bytes)
            if source_hash != str(next(
                row["sha256"] for row in json.loads(
                    zf.read(zip_member(zf, "StructuralGuard_handoff_20260903/02_excel_blind_inputs/blind_suite_manifest.json"))
                )["workbooks"] if row["file"] == source_name
            )):
                raise ValueError(f"source hash mismatch for {source_name}")
            specs = make_mutation_specs(entry)
            for condition, spec in (
                ("singleton", None),
                ("clean", None),
                ("coherent_block", specs[0]),
                ("systematic_column", specs[1]),
            ):
                case_id = f"sgv1_{condition}_{index:02d}"
                output_path = output_dir / "public" / "workbooks" / f"{case_id}.xlsx"
                case = build_case(source_bytes, entry, output_path, condition, spec)
                case["case_id"] = case_id
                case["source_sha256"] = source_hash
                case["source_seed"] = int(entry["seed"])
                cases.append(case)

    public_rows = [
        {
            "case_id": case["case_id"],
            "condition": case["condition"],
            "scenario": case["scenario"],
            "workbook": case["workbook"],
        }
        for case in cases
    ]
    with (output_dir / "public" / "manifest.csv").open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(public_rows[0]))
        writer.writeheader()
        writer.writerows(public_rows)
    labels = {
        "protocol": PROTOCOL,
        "source_archive": source_zip.name,
        "source_archive_sha256": archive_sha,
        "input_scope": "only the five workbooks in the source archive",
        "conditions": {
            "singleton": "shipped five workbooks with 28 isolated injected errors",
            "clean": "all 28 shipped errors repaired; zero injected errors",
            "coherent_block": "one fixed wrong formula pattern over 20 contiguous cells per workbook",
            "systematic_column": "one fixed wrong formula pattern over every formula cell in one column per workbook",
        },
        "cases": cases,
    }
    (output_dir / "labels" / "answer_keys.json").write_text(
        json.dumps(labels, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    metadata = {
        "protocol": PROTOCOL,
        "source_archive": source_zip.name,
        "source_archive_sha256": archive_sha,
        "case_count": len(cases),
        "conditions": {condition: sum(case["condition"] == condition for case in cases) for condition in ("singleton", "clean", "coherent_block", "systematic_column")},
        "formula_cells": sum(case["formula_cells"] for case in cases),
        "injected_errors": sum(len(case["errors"]) for case in cases),
        "generated_by": "scripts/build_structuralguard_standard_benchmark.py",
    }
    (output_dir / "metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    lines = []
    for path in sorted(output_dir.rglob("*")):
        if path.is_file() and path.name != "SHA256SUMS.txt":
            lines.append(f"{sha256_file(path)}  {path.relative_to(output_dir).as_posix()}")
    (output_dir / "SHA256SUMS.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(metadata, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
