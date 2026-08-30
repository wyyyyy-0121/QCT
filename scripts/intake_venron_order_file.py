#!/usr/bin/env python3
"""Read only VEnron's order metadata workbook and record its schema."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import tempfile
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Mapping, Sequence

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
PROTOCOL = "formulaguard_vhrl_venron_order_schema_v0"
INVENTORY_PROTOCOL = "formulaguard_vhrl_venron_inventory_v0"
ORDER_MEMBER = "VEnron1.0/Version/FileOrder.xls"
GROUP_PATTERN = re.compile(r"^(\d+)_(\d+)_(.+)$")
DEFAULT_ARCHIVE = ROOT / "data/external/model_discovery/raw/venron/VEnron1.0.7z"
DEFAULT_INVENTORY = ROOT / "results/venron_inventory_v0"
DEFAULT_DESTINATION = ROOT / "data/external/model_discovery/corpus/venron_order"
DEFAULT_OUTPUT = ROOT / "results/venron_order_schema_v0"
MAX_ORDER_FILE_BYTES = 64 * 1024 * 1024


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_json(path: Path, payload: Mapping[str, object]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n",
        encoding="ascii",
    )


def _git(command: Sequence[str], *, check: bool = True) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        ("git", *command),
        cwd=ROOT,
        check=check,
        capture_output=True,
        text=True,
    )


def require_clean_pushed_worktree() -> str:
    if _git(("status", "--porcelain", "--untracked-files=no")).stdout.strip():
        raise ValueError("tracked worktree must be clean before VEnron order intake")
    commit = _git(("rev-parse", "HEAD")).stdout.strip()
    upstream = _git(
        ("rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{upstream}")
    ).stdout.strip()
    pushed = _git(("merge-base", "--is-ancestor", commit, upstream), check=False)
    if pushed.returncode != 0:
        raise ValueError("VEnron order-intake implementation commit has not been pushed")
    return commit


def validate_layout(
    members: Sequence[Mapping[str, object]],
    *,
    expected_groups: int = 360,
    expected_group_workbooks: int = 7_294,
) -> dict[str, object]:
    workbook_paths = [
        PurePosixPath(str(row.get("member_path", "")))
        for row in members
        if row.get("kind") == "workbook"
    ]
    order_paths = [path for path in workbook_paths if path.as_posix() == ORDER_MEMBER]
    if len(order_paths) != 1:
        raise ValueError("VEnron inventory does not contain exactly one pinned order file")
    group_paths = [path for path in workbook_paths if path.as_posix() != ORDER_MEMBER]
    parents: dict[str, int] = Counter(path.parent.name for path in group_paths)
    if len(group_paths) != expected_group_workbooks or len(parents) != expected_groups:
        raise ValueError("VEnron group or workbook count differs from the frozen layout")

    group_ids: set[int] = set()
    for parent, observed_count in parents.items():
        match = GROUP_PATTERN.fullmatch(parent)
        if not match or int(match.group(2)) != observed_count:
            raise ValueError(f"VEnron group directory count is invalid: {parent!r}")
        group_ids.add(int(match.group(1)))
    if group_ids != set(range(1, expected_groups + 1)):
        raise ValueError("VEnron group IDs are not continuous and unique")
    return {
        "order_member": ORDER_MEMBER,
        "evolution_groups": len(parents),
        "group_workbooks": len(group_paths),
        "group_ids_sha256": hashlib.sha256(
            ",".join(str(value) for value in sorted(group_ids)).encode("ascii")
        ).hexdigest(),
    }


def resolve_executable(value: str, fallback: str) -> str:
    executable = shutil.which(value) if os.sep not in value else value
    executable = executable or shutil.which(fallback)
    if not executable or not Path(executable).is_file():
        raise ValueError(f"required executable not found: {value}")
    return str(Path(executable).resolve())


def extract_order_member(archive: Path, destination: Path, bsdtar: str) -> None:
    completed = subprocess.run(
        (bsdtar, "-xOf", str(archive), ORDER_MEMBER),
        check=True,
        capture_output=True,
    )
    if not completed.stdout or len(completed.stdout) > MAX_ORDER_FILE_BYTES:
        raise ValueError("VEnron order workbook is empty or exceeds the size limit")
    destination.write_bytes(completed.stdout)


def convert_order_workbook(source: Path, destination: Path, libreoffice: str) -> str:
    version = subprocess.run(
        (libreoffice, "--version"),
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()
    with tempfile.TemporaryDirectory(prefix="venron_order_lo_") as directory:
        temporary = Path(directory)
        profile = temporary / "profile"
        converted = temporary / "converted"
        profile.mkdir()
        converted.mkdir()
        subprocess.run(
            (
                libreoffice,
                f"-env:UserInstallation={profile.resolve().as_uri()}",
                "--headless",
                "--nologo",
                "--nodefault",
                "--nolockcheck",
                "--nofirststartwizard",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(converted),
                str(source),
            ),
            check=True,
            capture_output=True,
            timeout=120,
        )
        outputs = list(converted.glob("*.xlsx"))
        if len(outputs) != 1:
            raise ValueError("LibreOffice did not produce exactly one order workbook")
        shutil.copyfile(outputs[0], destination)
    return version


def _json_value(value: object) -> object:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        text = value if not isinstance(value, str) else value[:1024]
        return text
    return str(value)[:1024]


def inspect_order_workbook(path: Path, *, sample_rows: int = 25) -> dict[str, object]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, object]] = []
    try:
        for sheet in workbook.worksheets:
            types: Counter[str] = Counter()
            nonempty = 0
            samples: list[list[dict[str, object]]] = []
            for row in sheet.iter_rows():
                observed: list[dict[str, object]] = []
                for cell in row:
                    if cell.value is None:
                        continue
                    nonempty += 1
                    types[str(cell.data_type)] += 1
                    observed.append({
                        "column": cell.column,
                        "data_type": str(cell.data_type),
                        "value": _json_value(cell.value),
                    })
                if observed and len(samples) < sample_rows:
                    samples.append(observed)
            sheets.append({
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "nonempty_cells": nonempty,
                "cell_data_types": dict(sorted(types.items())),
                "sample_nonempty_rows": samples,
            })
    finally:
        workbook.close()
    if not sheets or sum(int(sheet["nonempty_cells"]) for sheet in sheets) == 0:
        raise ValueError("VEnron order workbook has no readable metadata cells")
    return {"sheet_count": len(sheets), "sheets": sheets}


def intake(
    *,
    archive: Path,
    inventory_dir: Path,
    destination: Path,
    output_dir: Path,
    bsdtar: str,
    libreoffice: str,
) -> Path:
    archive = archive.resolve()
    inventory_dir = inventory_dir.resolve()
    destination = destination.resolve()
    output_dir = output_dir.resolve()
    if destination.exists() or output_dir.exists():
        raise ValueError("completed or partial VEnron order intake output already exists")
    commit = require_clean_pushed_worktree()

    receipt_path = inventory_dir / "inventory_receipt.json"
    manifest_path = inventory_dir / "member_manifest.json"
    receipt = json.loads(receipt_path.read_text(encoding="ascii"))
    manifest = json.loads(manifest_path.read_text(encoding="ascii"))
    if (
        receipt.get("protocol") != INVENTORY_PROTOCOL
        or receipt.get("complete") is not True
        or receipt.get("archive_members_extracted") != 0
        or receipt.get("workbook_contents_read") != 0
        or receipt.get("protected_data_inputs") != []
        or receipt.get("member_manifest_sha256") != sha256(manifest_path)
        or manifest.get("protocol") != INVENTORY_PROTOCOL
        or manifest.get("archive_sha256") != receipt.get("archive_sha256")
        or sha256(archive) != receipt.get("archive_sha256")
    ):
        raise ValueError("VEnron inventory or archive identity changed")
    members = manifest.get("members")
    if not isinstance(members, list):
        raise ValueError("VEnron member manifest is malformed")
    layout = validate_layout(members)
    bsdtar_path = resolve_executable(bsdtar, "bsdtar")
    libreoffice_path = resolve_executable(libreoffice, "libreoffice")

    staging_destination = destination.with_name(f".{destination.name}.{os.getpid()}.tmp")
    staging_output = output_dir.with_name(f".{output_dir.name}.{os.getpid()}.tmp")
    if staging_destination.exists() or staging_output.exists():
        raise ValueError("VEnron order-intake staging path already exists")
    staging_destination.mkdir(parents=True)
    staging_output.mkdir(parents=True)
    installed_destination = False
    try:
        raw_order = staging_destination / "FileOrder.xls"
        converted_order = staging_destination / "FileOrder.xlsx"
        extract_order_member(archive, raw_order, bsdtar_path)
        libreoffice_version = convert_order_workbook(
            raw_order, converted_order, libreoffice_path
        )
        schema = inspect_order_workbook(converted_order)
        schema_path = staging_output / "order_schema.json"
        _write_json(schema_path, {"protocol": PROTOCOL, "schema": schema})
        result = {
            "protocol": PROTOCOL,
            "implementation_commit": commit,
            "archive_sha256": receipt["archive_sha256"],
            "inventory_receipt_sha256": sha256(receipt_path),
            "member_manifest_sha256": sha256(manifest_path),
            "layout": layout,
            "order_member": ORDER_MEMBER,
            "raw_order_sha256": sha256(raw_order),
            "converted_order_sha256": sha256(converted_order),
            "order_schema_sha256": sha256(schema_path),
            "bsdtar": bsdtar_path,
            "libreoffice": libreoffice_path,
            "libreoffice_version": libreoffice_version,
            "metadata_workbook_contents_read": 1,
            "evolution_group_workbook_contents_read": 0,
            "fault_label_inputs": [],
            "protected_data_inputs": [],
            "complete": True,
        }
        _write_json(staging_output / "order_intake_receipt.json", result)
        destination.parent.mkdir(parents=True, exist_ok=True)
        output_dir.parent.mkdir(parents=True, exist_ok=True)
        os.replace(staging_destination, destination)
        installed_destination = True
        os.replace(staging_output, output_dir)
    except BaseException:
        shutil.rmtree(staging_destination, ignore_errors=True)
        shutil.rmtree(staging_output, ignore_errors=True)
        if installed_destination:
            shutil.rmtree(destination, ignore_errors=True)
        raise
    return output_dir / "order_intake_receipt.json"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--archive", type=Path, default=DEFAULT_ARCHIVE)
    parser.add_argument("--inventory", type=Path, default=DEFAULT_INVENTORY)
    parser.add_argument("--destination", type=Path, default=DEFAULT_DESTINATION)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--bsdtar", default="bsdtar")
    parser.add_argument("--libreoffice", default="libreoffice")
    args = parser.parse_args()
    try:
        receipt = intake(
            archive=args.archive,
            inventory_dir=args.inventory,
            destination=args.destination,
            output_dir=args.output,
            bsdtar=args.bsdtar,
            libreoffice=args.libreoffice,
        )
    except (OSError, ValueError, KeyError, json.JSONDecodeError, subprocess.SubprocessError) as exc:
        raise SystemExit(f"VEnron order intake refused: {exc}") from exc
    print(receipt)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
