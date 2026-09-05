import csv
import importlib.util
import json
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import mock

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts/run_header_partition_predictions.py"
SPEC = importlib.util.spec_from_file_location("run_header_partition_predictions", SCRIPT)
assert SPEC is not None and SPEC.loader is not None
runner = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(runner)


GROUP_FIELDS = (
    "cohort",
    "workbook",
    "workbook_sha256",
    "structure_cluster_id",
)


def write_partition_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ledger"
    for column, value in enumerate(
        ("EAST", "EAST", "WEST", "WEST", "NORTH", "NORTH", "SOUTH", "SOUTH"),
        1,
    ):
        sheet.cell(3, column, value)
    for column, value in enumerate(
        ("TOTAL FLOW", "FLOW EAST", "FLOW WEST", "FLOW NORTH", "FLOW SOUTH"),
        10,
    ):
        sheet.cell(4, column, value)
    for row in range(6, 9):
        for column in range(1, 9):
            sheet.cell(row, column, row * 10 + column)
        sheet[f"J{row}"] = f"=SUM(A{row}:H{row})"
        sheet[f"K{row}"] = f"=SUM(A{row})"
        sheet[f"L{row}"] = f"=SUM(C{row}:D{row})"
        sheet[f"M{row}"] = f"=SUM(E{row}:F{row})"
        sheet[f"N{row}"] = f"=SUM(G{row}:H{row})"
    path.parent.mkdir(parents=True)
    workbook.save(path)


def write_groups(path: Path, rows: list[dict[str, str]], extra_field: str | None = None) -> None:
    fields = [*GROUP_FIELDS]
    if extra_field:
        fields.append(extra_field)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


class HeaderPartitionPredictionTests(unittest.TestCase):
    def test_run_is_label_free_deduplicated_and_byte_stable(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            digest = runner.sha256(workbook)
            groups = root / "groups.csv"
            row = {
                "cohort": "public:test",
                "workbook": "data/public/model.xlsx",
                "workbook_sha256": digest,
                "structure_cluster_id": "structure:test",
            }
            write_groups(groups, [row, row])

            first = root / "first"
            second = root / "second"
            runner.run(
                groups=groups,
                output=first,
                cohorts=("public:test",),
                workers=1,
                root=root,
                allowed_roots=(root / "data",),
                allowed_group_roots=(root,),
                allow_dirty=True,
            )
            runner.run(
                groups=groups,
                output=second,
                cohorts=("public:test",),
                workers=1,
                root=root,
                allowed_roots=(root / "data",),
                allowed_group_roots=(root,),
                allow_dirty=True,
            )

            for name in (
                "completion_receipt.json",
                "predictions.jsonl",
                "scan_summary.json",
            ):
                self.assertEqual((first / name).read_bytes(), (second / name).read_bytes())
            record = json.loads((first / "predictions.jsonl").read_text())
            self.assertEqual(record["cohort"], "public:test")
            self.assertEqual(record["qualified_block_count"], 1)
            self.assertEqual(record["label_inputs"], [])
            summary = json.loads((first / "scan_summary.json").read_text())
            self.assertEqual(summary["selected_identity_rows"], 2)
            self.assertEqual(summary["unique_observed_workbooks"], 1)
            self.assertEqual(summary["schema_version"], 2)
            self.assertEqual(summary["global"]["action_workbooks"], 0)
            self.assertEqual(summary["actions"], [])
            self.assertEqual(summary["global"]["review_candidate_workbooks"], 1)
            review = summary["block_reviews"][0]
            self.assertEqual(len(review["review_cells"]), 3)
            self.assertEqual(
                review["review_cells"][0]["candidate_formula"], "=SUM(A6:B6)"
            )
            self.assertFalse(review["within_block_ranking_supported"])
            self.assertFalse(review["automatic_edit_supported"])
            self.assertFalse(review["can_identify_formula_error"])
            self.assertEqual(summary["label_inputs"], [])
            self.assertEqual(summary["protected_data_inputs"], [])
            self.assertTrue(summary["formal_evidence"])

    def test_label_columns_and_protected_paths_are_rejected(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            row = {
                "cohort": "public:test",
                "workbook": "data/public/model.xlsx",
                "workbook_sha256": runner.sha256(workbook),
                "structure_cluster_id": "structure:test",
                "source_cells": "Ledger!K6",
            }
            groups = root / "groups.csv"
            write_groups(groups, [row], extra_field="source_cells")
            with self.assertRaisesRegex(ValueError, "possible labels"):
                runner.load_units(
                    groups,
                    cohorts=("public:test",),
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root,),
                )

            row.pop("source_cells")
            row["workbook"] = "data/external/v5_psl/final_blind/model.xlsx"
            groups = root / "protected.csv"
            write_groups(groups, [row])
            with self.assertRaisesRegex(ValueError, "protected"):
                runner.load_units(
                    groups,
                    cohorts=("public:test",),
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root,),
                )

    def test_unknown_label_like_fields_are_rejected_by_exact_allowlist(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            base = {
                "cohort": "public:test",
                "workbook": "data/public/model.xlsx",
                "workbook_sha256": runner.sha256(workbook),
                "structure_cluster_id": "structure:test",
            }
            for field in ("ground_truth_formula", "is_faulty", "future_metadata"):
                groups = root / f"{field}.csv"
                write_groups(groups, [{**base, field: "hidden"}], extra_field=field)
                with self.subTest(field=field), self.assertRaisesRegex(
                    ValueError, "undeclared fields"
                ):
                    runner.load_units(
                        groups,
                        cohorts=("public:test",),
                        root=root,
                        allowed_roots=(root / "data",),
                        allowed_group_roots=(root,),
                    )

    def test_rows_with_extra_columns_are_rejected(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            groups = root / "groups.csv"
            groups.write_text(
                ",".join(GROUP_FIELDS)
                + "\n"
                + ",".join(
                    (
                        "public:test",
                        "data/public/model.xlsx",
                        runner.sha256(workbook),
                        "structure:test",
                        "unexpected-extra-value",
                    )
                )
                + "\n",
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ValueError, r"extra columns: \[None\]"):
                runner.load_units(
                    groups,
                    cohorts=("public:test",),
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root,),
                )

    def test_groups_path_suffix_protection_and_symlinks_are_rejected(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            row = {
                "cohort": "public:test",
                "workbook": "data/public/model.xlsx",
                "workbook_sha256": runner.sha256(workbook),
                "structure_cluster_id": "structure:test",
            }
            protected = root / "data/external/v5_psl/final_blind/groups.csv"
            protected.parent.mkdir(parents=True)
            write_groups(protected, [row])
            with self.assertRaisesRegex(ValueError, "protected"):
                runner.load_units(
                    protected,
                    cohorts=("public:test",),
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root / "data",),
                )

            wrong_suffix = root / "groups.txt"
            write_groups(wrong_suffix, [row])
            with self.assertRaisesRegex(ValueError, r"\.csv"):
                runner.load_units(
                    wrong_suffix,
                    cohorts=("public:test",),
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root,),
                )

            real_groups = root / "groups.csv"
            write_groups(real_groups, [row])
            linked_groups = root / "linked.csv"
            linked_groups.symlink_to(real_groups)
            with self.assertRaisesRegex(ValueError, "symlinked"):
                runner.load_units(
                    linked_groups,
                    cohorts=("public:test",),
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root,),
                )

            linked_workbook = root / "data/public/linked.xlsx"
            linked_workbook.symlink_to(workbook)
            linked_row = {
                **row,
                "workbook": "data/public/linked.xlsx",
            }
            linked_manifest = root / "linked-workbook.csv"
            write_groups(linked_manifest, [linked_row])
            with self.assertRaisesRegex(ValueError, "symlinked"):
                runner.load_units(
                    linked_manifest,
                    cohorts=("public:test",),
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root,),
                )

    def test_worker_parses_the_hash_verified_snapshot_not_mutated_source(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            expected_hash = runner.sha256(workbook)
            groups = root / "groups.csv"
            write_groups(
                groups,
                [
                    {
                        "cohort": "public:test",
                        "workbook": "data/public/model.xlsx",
                        "workbook_sha256": expected_hash,
                        "structure_cluster_id": "structure:test",
                    }
                ],
            )
            units, _ = runner.load_units(
                groups,
                cohorts=("public:test",),
                root=root,
                allowed_roots=(root / "data",),
                allowed_group_roots=(root,),
                snapshot_root=root / "snapshots",
            )
            workbook.write_bytes(b"changed after snapshot")

            unit = units[0]
            record = runner.predict_unit(
                (
                    {key: value for key, value in unit.items() if not key.startswith("_")},
                    unit["_snapshot_path"],
                )
            )

            self.assertEqual(record["workbook_sha256"], expected_hash)
            self.assertEqual(record["qualified_block_count"], 1)

    def test_dirty_sources_require_explicit_nonformal_mode(self):
        with mock.patch.object(
            runner, "_git_source_status", return_value=(" M formulaguard/workbook.py",)
        ):
            with self.assertRaisesRegex(ValueError, "clean tracked source"):
                runner.capture_source_state(ROOT)
            state = runner.capture_source_state(ROOT, allow_dirty=True)
        self.assertTrue(state["source_tree_dirty"])
        self.assertFalse(state["formal_evidence"])

    def test_output_cannot_overlap_groups_or_workbook_inputs(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            groups = root / "groups.csv"
            write_groups(
                groups,
                [
                    {
                        "cohort": "public:test",
                        "workbook": "data/public/model.xlsx",
                        "workbook_sha256": runner.sha256(workbook),
                        "structure_cluster_id": "structure:test",
                    }
                ],
            )
            with self.assertRaisesRegex(ValueError, "overlaps"):
                runner.run(
                    groups=groups,
                    output=root / "data",
                    cohorts=("public:test",),
                    workers=1,
                    root=root,
                    allowed_roots=(root / "data",),
                    allowed_group_roots=(root,),
                    allow_dirty=True,
                )

    def test_real_scoring_group_provenance_columns_are_explicitly_allowed(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "data/public/model.xlsx"
            write_partition_workbook(workbook)
            groups = root / "groups.csv"
            fields = [*GROUP_FIELDS, *runner.PROVENANCE_FIELDS_ALLOWED_BUT_NOT_READ]
            row = {
                "cohort": "public:test",
                "workbook": "data/public/model.xlsx",
                "workbook_sha256": runner.sha256(workbook),
                "structure_cluster_id": "structure:test",
                "cohort_instance_id": "public:test::1",
                "instance_id": "case-1",
                "provenance_group_id": "source-1",
                "outer_group_id": "outer-1",
            }
            with groups.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fields)
                writer.writeheader()
                writer.writerow(row)

            units, audit = runner.load_units(
                groups,
                cohorts=("public:test",),
                root=root,
                allowed_roots=(root / "data",),
                allowed_group_roots=(root,),
            )

            self.assertEqual(len(units), 1)
            self.assertEqual(
                set(audit["strict_scoring_group_field_allowlist"]),
                runner.ALLOWED_FIELDS,
            )


if __name__ == "__main__":
    unittest.main()
