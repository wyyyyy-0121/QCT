import tempfile
import unittest
from pathlib import Path

import openpyxl

from formulaguard.venron import (
    ORDER_MEMBER,
    compare_formula_profiles,
    inspect_formula_workbook,
    parse_order_workbook,
)
from scripts.prepare_venron_v0_corpus import validate_verbose_listing
from scripts.prepare_venron_v0_corpus import _hash_record
from scripts.score_venron_v0_gate import evaluate_gates


def profile(entries: list[tuple[str, str, str]]) -> dict[str, object]:
    return {
        "formulas": [
            {"sheet": sheet, "address": address, "formula": formula}
            for sheet, address, formula in entries
        ]
    }


class VEnronV0Tests(unittest.TestCase):
    def test_source_hash_keeps_publisher_and_archive_md5_separate(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            relative = "VEnron1.0/1_1_alpha/v1.xls"
            source = root / relative
            source.parent.mkdir(parents=True)
            source.write_bytes(b"archive member bytes")
            result = _hash_record(({
                "source_relative_path": relative,
                "source_md5": "a" * 32,
            }, str(root)))
            self.assertFalse(result["publisher_md5_matches_archive_bytes"])
            self.assertEqual(len(result["archive_member_md5"]), 32)
            self.assertEqual(len(result["source_sha256"]), 64)

    def test_archive_types_reject_links_and_devices(self):
        self.assertEqual(
            validate_verbose_listing("drwx group/\n-rw- group/file.xls\n", 2),
            {"-": 1, "d": 1},
        )
        with self.assertRaisesRegex(ValueError, "unsupported member types"):
            validate_verbose_listing("lrwx group/link\n", 1)

    def test_order_parser_uses_author_order_md5_and_path(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "order.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Group List"
            sheet = workbook.create_sheet("1_2_a_name_that_is_truncated")
            full_group = "1_2_a_name_that_is_truncated_in_the_sheet_title"
            md5_a_token = "a" * 31
            md5_a = "0" + md5_a_token
            md5_b = "b" * 32
            file_a = f"v1.xls_2001-01-01-00-00_{md5_a_token}.xls"
            file_b = f"v2.xls_2001-01-02-00-00_{md5_b}.xls"
            sheet.append(["1", "File Name", "Sending Time", "Sender", "File MD5", "File Path"])
            sheet.append([2, "v2.xls", None, "ignored", md5_b, f"..\\{full_group}\\{file_b}"])
            sheet.append([None, 1, "email fields are ignored", None, None, None])
            sheet.append([1, "v1.xls", None, "ignored", md5_a_token, f"..\\{full_group}\\{file_a}"])
            workbook.save(path)
            workbook.close()
            members = {
                ORDER_MEMBER,
                f"VEnron1.0/{full_group}/{file_a}",
                f"VEnron1.0/{full_group}/{file_b}",
            }
            rows = parse_order_workbook(
                path,
                members,
                expected_groups=1,
                expected_workbooks=2,
            )
            self.assertEqual([row["version_order"] for row in rows], [1, 2])
            self.assertEqual([row["source_md5"] for row in rows], [md5_a, md5_b])
            self.assertEqual(rows[0]["source_md5_token"], md5_a_token)
            self.assertEqual(rows[0]["group_name"], full_group)

    def test_formula_inspection_exports_formulas_not_constants(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "book.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["A1"] = 10
            sheet["A2"] = "private constant"
            sheet["A3"] = "=SUM(A1:A2)"
            workbook.save(path)
            workbook.close()
            observed = inspect_formula_workbook(path)
            self.assertEqual(observed["formula_count"], 1)
            self.assertEqual(observed["formulas"], [{
                "sheet": "Sheet",
                "address": "A3",
                "formula": "=SUM(A1:A2)",
            }])
            self.assertNotIn("private constant", str(observed))

    def test_transition_separates_direct_move_bulk_and_no_change(self):
        single = compare_formula_profiles(
            profile([("S", "A1", "=1"), ("S", "A2", "=2")]),
            profile([("S", "A1", "=3"), ("S", "A2", "=2")]),
        )
        self.assertTrue(single["single_direct"])
        self.assertFalse(single["nonbulk_multi_direct"])

        moved = compare_formula_profiles(
            profile([("S", "A1", "=B1")]),
            profile([("S", "A2", "=B1")]),
        )
        self.assertTrue(moved["address_only_formula_move"])
        self.assertEqual(moved["direct_formula_text_changes"], 0)

        left = [("S", f"A{index}", f"={index}") for index in range(1, 41)]
        right = [("S", f"A{index}", f"={index + 1}") for index in range(1, 41)]
        bulk = compare_formula_profiles(profile(left), profile(right))
        self.assertTrue(bulk["bulk_direct_rewrite"])
        self.assertFalse(bulk["nonbulk_multi_direct"])

        unchanged = compare_formula_profiles(profile(left), profile(left))
        self.assertTrue(unchanged["no_formula_text_change"])

    def test_gate_requires_nonbulk_multi_edit_coverage(self):
        summary = {
            "evolution_groups": 360,
            "source_workbooks": 7294,
            "direct_change_transitions": 1000,
            "direct_change_groups": 150,
            "nonbulk_multi_direct_transitions": 99,
            "nonbulk_multi_direct_groups": 30,
            "parse_coverage": 0.90,
            "single_direct_transitions": 1,
            "multi_direct_transitions": 1,
            "bulk_direct_rewrite_transitions": 1,
            "bulk_add_remove_transitions": 1,
            "address_only_formula_move_transitions": 1,
            "formula_addition_transitions": 1,
            "formula_removal_transitions": 1,
            "no_formula_text_change_transitions": 1,
            "ineligible_adjacent_transitions": 1,
            "fault_label_inputs": [],
            "protected_data_inputs": [],
            "cached_value_difference_inputs": [],
            "v4_inputs": [],
        }
        gates = evaluate_gates(summary)
        self.assertFalse(gates["nonbulk_multi_direct_100_transitions_across_30_groups"])
        summary["nonbulk_multi_direct_transitions"] = 100
        self.assertTrue(all(evaluate_gates(summary).values()))


if __name__ == "__main__":
    unittest.main()
