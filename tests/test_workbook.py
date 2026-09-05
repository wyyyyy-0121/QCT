import unittest
import xml.etree.ElementTree as ET
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from formulaguard.workbook import SharedFormulaRegion, WorkbookModel

WORKSHEET_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)


def _write_shared_formula_workbook(
    path: Path,
    *,
    members: tuple[str, ...] = ("C2", "C3", "C4"),
    master: str | None = "C2",
    master_ref: str | None = "C2:C4",
    shared_index: str = "7",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model"
    for address in members:
        row = int(address[1:])
        sheet[address] = f"=A{row}+B{row}"
    workbook.save(path)

    with ZipFile(path, "r") as source:
        archive_members = [
            (item, source.read(item.filename)) for item in source.infolist()
        ]
    worksheet_name = "xl/worksheets/sheet1.xml"
    worksheet = ET.fromstring(
        next(
            data
            for item, data in archive_members
            if item.filename == worksheet_name
        )
    )
    for address in members:
        cell = worksheet.find(
            f".//{{{WORKSHEET_NAMESPACE}}}c[@r='{address}']"
        )
        if cell is None:
            raise AssertionError(f"Missing test cell {address}")
        formula = cell.find(f"{{{WORKSHEET_NAMESPACE}}}f")
        if formula is None:
            raise AssertionError(f"Missing test formula {address}")
        formula.attrib.update({"t": "shared", "si": shared_index})
        if address == master:
            if master_ref is not None:
                formula.attrib["ref"] = master_ref
        else:
            formula.text = None

    rewritten = ET.tostring(
        worksheet,
        encoding="utf-8",
        xml_declaration=True,
    )
    temporary = path.with_suffix(".rewritten.xlsx")
    with ZipFile(temporary, "w") as destination:
        for item, data in archive_members:
            destination.writestr(
                item,
                rewritten if item.filename == worksheet_name else data,
            )
    temporary.replace(path)


def simple_model():
    cells = {
        ("Model", "A1"): 2,
        ("Model", "B1"): 3,
        ("Model", "A2"): 4,
        ("Model", "B2"): 5,
    }
    formulas = {
        ("Model", "C1"): "=A1+B1",
        ("Model", "C2"): "=A2+B2",
        ("Model", "D1"): "=C1*2",
        ("Model", "D2"): "=SUM(C1:C2)",
        ("Model", "E1"): "=IF(D2>10,D2,0)",
    }
    return WorkbookModel.from_cells(cells, formulas)


class WorkbookTests(unittest.TestCase):
    def test_from_cells_requires_explicit_metadata_completeness(self):
        self.assertFalse(simple_model().header_partition_metadata_complete)

        model = WorkbookModel.from_cells(
            {("Model", "A1"): "heading", ("Hidden", "A1"): "secret"},
            {("Model", "C2"): "=SUM(A2:B2)"},
            cell_visibility={("Model", "A1"): False},
            number_formats={("Model", "C2"): "0.00"},
            sheet_visibility={"Model": True, "Hidden": False},
            merged_ranges={"Model": (("D1", "E1"),)},
            formula_kinds={("Model", "C2"): "shared"},
            formula_regions={"Model": (("F2", "G3", "array"),)},
            shared_formula_groups={("Model", "C2"): "Model:7"},
            hidden_rows={"Model": (4,)},
            hidden_columns={"Model": ((8, 9),)},
            header_partition_metadata_complete=True,
        )

        self.assertTrue(model.header_partition_metadata_complete)
        self.assertFalse(model.is_visible(("Model", "A1")))
        self.assertFalse(model.is_visible(("Hidden", "A1")))
        self.assertFalse(model.is_visible(("Model", "D4")))
        self.assertFalse(model.is_visible(("Model", "H100")))
        self.assertTrue(model.is_visible(("Model", "G100")))
        self.assertEqual(model.number_format(("Model", "C2")), "0.00")
        self.assertTrue(model.is_merged(("Model", "E1")))
        self.assertEqual(model.formula_kind(("Model", "C2")), "shared")
        self.assertEqual(
            model.shared_formula_group(("Model", "C2")),
            "Model:7",
        )
        self.assertEqual(model.formula_kind(("Model", "G3")), "array")
        self.assertTrue(model.is_formula_derived(("Model", "F3")))

    def test_xlsx_reader_tracks_shared_formula_provenance(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "shared.xlsx"
            _write_shared_formula_workbook(path)

            model = WorkbookModel.from_xlsx(path)

        self.assertTrue(model.header_partition_metadata_complete)
        self.assertEqual(model.formulas[("Model", "C2")], "=A2+B2")
        self.assertEqual(model.formulas[("Model", "C3")], "=A3+B3")
        self.assertEqual(model.formulas[("Model", "C4")], "=A4+B4")
        groups = {
            model.shared_formula_group(("Model", f"C{row}"))
            for row in range(2, 5)
        }
        self.assertEqual(groups, {"Model:7"})
        self.assertTrue(
            all(
                model.formula_kind(("Model", f"C{row}")) == "shared"
                for row in range(2, 5)
            )
        )
        self.assertEqual(
            model.shared_formula_regions,
            (
                SharedFormulaRegion(
                    sheet="Model",
                    group_id="Model:7",
                    master_cell=("Model", "C2"),
                    start="C2",
                    end="C4",
                    master_formula="=A2+B2",
                    members=(("Model", "C2"), ("Model", "C3"), ("Model", "C4")),
                ),
            ),
        )

    def test_xlsx_reader_rejects_incomplete_shared_formula_provenance(self):
        cases = {
            "missing master ref": {
                "master_ref": None,
            },
            "malformed master ref": {
                "master_ref": "C2::C4",
            },
            "master outside ref": {
                "master_ref": "C3:C5",
            },
            "follower outside ref": {
                "master_ref": "C2:C3",
            },
            "sparse group": {
                "members": ("C2", "C4"),
            },
            "missing master": {
                "master": None,
            },
            "malformed shared index": {
                "shared_index": "not-an-index",
            },
        }
        with TemporaryDirectory() as directory:
            for name, options in cases.items():
                with self.subTest(name=name):
                    path = Path(directory) / f"{name.replace(' ', '-')}.xlsx"
                    _write_shared_formula_workbook(path, **options)

                    model = WorkbookModel.from_xlsx(path)

                    self.assertFalse(
                        model.header_partition_metadata_complete
                    )
                    if name == "missing master ref":
                        self.assertEqual(
                            {
                                model.shared_formula_group(
                                    ("Model", address)
                                )
                                for address in ("C2", "C3", "C4")
                            },
                            {"Model:7"},
                        )
                    if name == "sparse group":
                        self.assertEqual(len(model.shared_formula_regions), 1)
                        region = model.shared_formula_regions[0]
                        self.assertEqual((region.start, region.end), ("C2", "C4"))
                        self.assertEqual(
                            region.members,
                            (("Model", "C2"), ("Model", "C4")),
                        )
                    else:
                        self.assertEqual(model.shared_formula_regions, ())

    def test_xlsx_reader_preserves_header_partition_safety_metadata(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "metadata.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Model"
            sheet.merge_cells("A1:B1")
            sheet["A1"] = "Merged heading"
            sheet["C2"] = "=SUM(A2:B2)"
            sheet["E2"] = ArrayFormula(
                ref="E2:E4", text="=SUM(A2:A4*B2:B4)"
            )
            sheet["G2"] = DataTableFormula(
                ref="G2:H4", dt2D=True, r1="A1", r2="B1"
            )
            sheet["L2"] = ArrayFormula(ref="L2", text="=SUM(A2:B2)")
            sheet.row_dimensions[7].hidden = True
            sheet.column_dimensions["J"].hidden = True
            workbook.save(path)

            model = WorkbookModel.from_xlsx(path)

        self.assertTrue(model.header_partition_metadata_complete)
        self.assertEqual(model.merged_ranges, {"Model": (("A1", "B1"),)})
        self.assertTrue(model.is_merged(("Model", "A1")))
        self.assertTrue(model.is_merged(("Model", "B1")))
        self.assertFalse(model.is_merged(("Model", "C1")))
        self.assertEqual(model.formula_kind(("Model", "C2")), "normal")
        self.assertEqual(model.formula_kind(("Model", "E2")), "array")
        self.assertEqual(model.formula_kind(("Model", "E4")), "array")
        self.assertEqual(model.formula_kind(("Model", "G3")), "dataTable")
        self.assertEqual(model.formula_kind(("Model", "L2")), "array")
        self.assertTrue(model.is_formula_derived(("Model", "E3")))
        self.assertTrue(model.is_formula_derived(("Model", "H4")))
        self.assertFalse(model.is_formula_derived(("Model", "H5")))
        self.assertFalse(model.is_visible(("Model", "A7")))
        self.assertFalse(model.is_visible(("Model", "J100")))
        self.assertTrue(model.is_visible(("Model", "K100")))

    def test_supported_formula_evaluation(self):
        values, errors = simple_model().evaluate()
        self.assertFalse(errors)
        self.assertAlmostEqual(values[("Model", "C1")], 5)
        self.assertAlmostEqual(values[("Model", "D2")], 14)
        self.assertAlmostEqual(values[("Model", "E1")], 14)

    def test_dependency_graph_and_propagation_depth(self):
        model = simple_model()
        graph = model.dependency_graph()
        self.assertIn(("Model", "C1"), graph.dependents[("Model", "A1")])
        self.assertIn(("Model", "E1"), graph.descendants(("Model", "C1")))
        self.assertEqual(graph.shortest_path_length(("Model", "C1"), ("Model", "E1")), 2)
        self.assertEqual(graph.shortest_path(("Model", "C1"), ("Model", "E1")), [
            ("Model", "C1"),
            ("Model", "D2"),
            ("Model", "E1"),
        ])

    def test_counterfactual_override_changes_downstream_value(self):
        model = simple_model()
        base, base_errors = model.evaluate()
        changed, changed_errors = model.evaluate({("Model", "C1"): "=A1-B1"})
        self.assertFalse(base_errors)
        self.assertFalse(changed_errors)
        self.assertNotEqual(base[("Model", "E1")], changed[("Model", "E1")])

    def test_value_override_changes_inputs_without_mutating_model(self):
        model = simple_model()
        base, base_errors = model.evaluate()
        changed, changed_errors = model.evaluate(value_overrides={("Model", "A1"): 20})
        repeated, repeated_errors = model.evaluate()
        self.assertFalse(base_errors)
        self.assertFalse(changed_errors)
        self.assertFalse(repeated_errors)
        self.assertNotEqual(base[("Model", "C1")], changed[("Model", "C1")])
        self.assertEqual(base, repeated)

    def test_value_override_rejects_formula_cell(self):
        model = simple_model()
        with self.assertRaisesRegex(ValueError, "formula cells"):
            model.evaluate(value_overrides={("Model", "C1"): 99})

    def test_if_only_evaluates_the_selected_branch(self):
        model = WorkbookModel.from_cells(
            {("Model", "A1"): 1, ("Model", "B1"): 0},
            {
                ("Model", "C1"): "=IF(A1>0,1,1/B1)",
                ("Model", "C2"): "=IF(A1<0,1/B1,2)",
            },
        )

        values, errors = model.evaluate()

        self.assertFalse(errors)
        self.assertEqual(values[("Model", "C1")], 1)
        self.assertEqual(values[("Model", "C2")], 2)

    def test_targeted_evaluation_matches_full_result_and_skips_unrelated_formulas(self):
        model = WorkbookModel.from_cells(
            {("Model", "A1"): 2, ("Model", "Z2"): 0},
            {
                ("Model", "B1"): "=A1+1",
                ("Model", "C1"): "=B1*2",
                ("Model", "Z1"): "=1/Z2",
            },
        )

        full_values, full_errors = model.evaluate()
        values, errors = model.evaluate(targets=[("Model", "C1")])

        self.assertEqual(values[("Model", "C1")], full_values[("Model", "C1")])
        self.assertEqual(values[("Model", "B1")], full_values[("Model", "B1")])
        self.assertNotIn(("Model", "Z1"), values)
        self.assertNotIn(("Model", "Z1"), errors)
        self.assertIn(("Model", "Z1"), full_errors)

    def test_full_column_range_materializes_only_present_cells(self):
        model = WorkbookModel.from_cells(
            {("Model", "A1"): 10, ("Model", "A3"): 20},
            {("Model", "B1"): "=AVERAGE(A1:A1048576)"},
        )

        graph = model.dependency_graph()
        values, errors = model.evaluate()
        changed, changed_errors = model.evaluate(
            value_overrides={("Model", "A2"): 30}
        )

        self.assertEqual(
            graph.precedents[("Model", "B1")],
            {("Model", "A1"), ("Model", "A3")},
        )
        self.assertFalse(errors)
        self.assertFalse(changed_errors)
        self.assertEqual(values[("Model", "B1")], 15)
        self.assertEqual(changed[("Model", "B1")], 20)


if __name__ == "__main__":
    unittest.main()
